# src/water_validation/checks.py
from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import re
import math

#from llm_client import classify_funding_with_confidence
# from .llm_client import classify_funding_with_confidence


#from models import CheckResult, Severity, Status
from .models import CheckResult, Severity, Status

from difflib import SequenceMatcher
from .prompts import SUBJECT_TO_ALLOWED, SUBJECT_SYNONYMS, ALLOWED_FUNDING_LABELS, build_llm_prompt


#from config import PlanConfig
from .config import PlanConfig

#from models import CheckResult, Severity, Status
from .models import CheckResult, Severity, Status

from .utils import (
    detect_city_cols,
    excel_row_to_df_index,
    normalize_text,
    parse_ratio,
    round_half_up,
)

LOCATION_HELPER_TEXTS = {'רחוב/שכונה/תב"ע', 'רחוב/שכונה/תב״ע'}


# =============================================================================
# Local helpers (kept here to keep checks self-contained)
# =============================================================================
from pathlib import Path

def print_rule_kpi(rule_id: str, results: List[CheckResult], file_path: str | None = None) -> None:
    """
    Prints only:
      <file.xlsx> | <rule_id> | <fail>/<total> FAIL
    """
    file_name = Path(file_path).name if file_path else "-"

    if not results:
        print(f"{file_name} | {rule_id} | 0/0 FAIL")
        return

    row_idxs = [r.row_index for r in results if getattr(r, "row_index", None) is not None]
    if row_idxs:
        total = len(set(row_idxs))
        fails = len(set(
            r.row_index for r in results
            if getattr(r, "status", None) == Status.FAIL and getattr(r, "row_index", None) is not None
        ))
    else:
        total = len(results)
        fails = sum(1 for r in results if getattr(r, "status", None) == Status.FAIL)

    print(f"{file_name} | {rule_id} | {fails}/{total} FAIL")




def fmt_num(x, decimals=3):
    """Format numbers nicely (thousands separators, trim trailing zeros)."""
    try:
        v = float(x)
    except Exception:
        return str(x)
    s = f"{v:,.{decimals}f}"
    return s.rstrip("0").rstrip(".")



def _col_idx_to_letter(col_idx: int) -> str:
    """Convert 0-based column index to Excel column letter(s). 0=A, 17=R, 26=AA."""
    if col_idx < 0:
        return ""
    if col_idx < 26:
        return chr(65 + col_idx)
    q, r = divmod(col_idx, 26)
    return chr(64 + q) + chr(65 + r)


def _summary_sheet_cell(cfg: PlanConfig, excel_row_1based: int, col_letter: str) -> str:
    """Exact Excel coordinate for the Summary sheet, e.g. 'סיכום תכנית השקעות!R15'."""
    return f"{cfg.sheet_name}!{col_letter}{excel_row_1based}"


def get_cell(plan_df: pd.DataFrame, df_row_idx: int, col_idx: int) -> Any:
    """
    Safe cell getter (by integer positions).
    Returns None if out-of-bounds instead of raising IndexError.
    """
    if df_row_idx < 0 or df_row_idx >= len(plan_df):
        return None
    if col_idx < 0 or col_idx >= plan_df.shape[1]:
        return None
    return plan_df.iat[df_row_idx, col_idx]


def is_non_empty(value: Any) -> bool:
    return pd.notna(value) and str(value).strip() != ""


def fail_no_cities(rule_id: str, rule_name: str, cfg: PlanConfig) -> List[CheckResult]:
    return [
        CheckResult(
            rule_id=rule_id,
            rule_name=rule_name,
            severity=Severity.CRITICAL,
            sheet_name=cfg.sheet_name,
            row_index=None,
            column_name=None,
            key_context="city_cols_detection",
            actual_value=None,
            expected_value="at least 1 city column",
            status=Status.FAIL,
            message="No city columns detected (header row selection or heuristics mismatch).",
        )
    ]


import json

def load_kinun_store(json_path: str) -> dict:
    """
    Load kinun JSON. Returns dict with keys:
      - "utilities": {utility_name: {water_full, water_reduced, sewer_full, sewer_reduced}, ...}
      - "year": baseline year from JSON (e.g. 2026), for use in R_1 messages.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {
        "utilities": data["utilities"],
        "year": data.get("year"),
    }

def lookup_kinun_value(kinun_store: dict, utility: str, col_name: str) -> float:
    """
    kinun_store: dict like {"הרי נצרת": {"water_full": 201993, ...}, ...}
    col_name: one of: water_full, water_reduced, sewer_full, sewer_reduced
    """
    u = str(utility).strip()
    if u not in kinun_store:
        raise KeyError(f"Utility not found in kinun JSON: {u}")

    val = kinun_store[u].get(col_name)
    if val is None:
        raise KeyError(f"Missing kinun value: utility={u}, col_name={col_name}")

    return float(val)



# =============================================================================
# Checks
# =============================================================================


# Epsilon for R_1 kinun value comparison (allow insignificant rounding differences)
R1_KINUN_EPSILON = 1.0


def check_001_kinun_values_rounded(
    plan_df: pd.DataFrame,
    kinun_store,
    utility: str,
    cfg: PlanConfig,
    kinun_year: Optional[int] = None,
) -> List[CheckResult]:
    """
    Compare plan kinun values vs kinun reference (e.g. kinun_values_2026.json).
    Plan values are read from column R at fixed Excel rows 8, 9, 11, 12.
    Mapping: R8=Total Plan (Col B), R9=Sewage (Col D), R11=Water (Col C), R12=Reclaimed/Other (Col E).
    Uses epsilon=1.0 for float tolerance.
    If the plan was filled with a different year's kinun (e.g. 2024), use that year's JSON or update the plan.
    """
    # Explicit mapping: (Excel row, JSON key, report label for messages, מפתח בדיקה display)
    # Excel Row 8 (Water Full) -> water_full | ערכי כינון מים מלא
    # Excel Row 9 (Sewage Full) -> sewer_full | ערכי כינון ביוב מלא
    # Excel Row 11 (Water Reduced) -> water_reduced | ערכי כינון מים מופחת
    # Excel Row 12 (Sewage Reduced) -> sewer_reduced | ערך כינון ביוב מופחת
    ROW_JSON_LABEL = [
        (8, "water_full", "Water Full Value", "ערכי כינון מים מלא"),
        (9, "sewer_full", "Sewage Full Value", "ערכי כינון ביוב מלא"),
        (11, "water_reduced", "Water Reduced Value", "ערכי כינון מים מופחת"),
        (12, "sewer_reduced", "Sewage Reduced Value", "ערך כינון ביוב מופחת"),
    ]

    results: List[CheckResult] = []

    for excel_row, json_key, report_label, rule_key_hebrew in ROW_JSON_LABEL:
        df_idx = excel_row_to_df_index(excel_row, cfg)
        plan_raw = get_cell(plan_df, df_idx, cfg.value_col_r_idx)
        kinun_raw = lookup_kinun_value(kinun_store, utility, json_key)

        plan_val = float(plan_raw) if plan_raw is not None and pd.notna(plan_raw) else 0.0
        kinun_val = float(kinun_raw)
        plan_round = round_half_up(plan_val, 0)
        kinun_round = round_half_up(kinun_val, 0)

        # Allow tiny floating-point tolerance (epsilon = 1.0)
        ok = abs(plan_round - kinun_round) <= R1_KINUN_EPSILON
        cell_ref = f"R{excel_row}"
        excel_cell = _summary_sheet_cell(cfg, excel_row, "R")

        year_label = f"ערכי כינון {kinun_year}" if kinun_year is not None else "ערכי כינון"
        if ok:
            message = (
                f"התאמה בין {rule_key_hebrew} בתכנית ({cell_ref}) לבין ערכי הכינון: "
                f"בתכנית ‏{fmt_num(plan_round, 0)}, ב{year_label} ‏{fmt_num(kinun_round, 0)}"
            )
        else:
            message = (
                f"חוסר התאמה ב{rule_key_hebrew} ({cell_ref}): בתכנית ‏{fmt_num(plan_round, 0)}, "
                f"ב{year_label} ‏{fmt_num(kinun_round, 0)}. יש לבדוק את הנתונים או לעדכן את קובץ ערכי הכינון לפי שנת התכנית."
            )

        # Rule id: R_1_<hebrew> so grouping stays R_1; מפתח בדיקה shows Hebrew name in export
        results.append(
            CheckResult(
                rule_id=f"R_1_{rule_key_hebrew}",
                rule_name="בדיקת ערכי כינון (עיגול לפני השוואה)",
                severity=Severity.CRITICAL,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name="R",
                key_context=f"plan_cell=R{excel_row}; kinun_col={json_key}",
                actual_value=plan_round,
                expected_value=kinun_round,
                status=Status.PASS_ if ok else Status.FAIL,
                message=message,
                excel_cells=[excel_cell],
            )
        )

    return results


# -----------------------------------------


def _row_label_from_col_b(plan_df: pd.DataFrame, df_idx: int, cfg: PlanConfig) -> str:
    """Component name from Column B (index 1); fallback to Column A (index 0)."""
    for col_idx in (cfg.data_marker_col_idx, cfg.label_col_idx):
        v = get_cell(plan_df, df_idx, col_idx)
        if v is not None and not (isinstance(v, float) and pd.isna(v)):
            s = str(v).strip()
            if s and s.lower() != "nan":
                return s
    return "שורה " + str(df_idx)


def check_002_asset_ratio(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    Rule R_2: יחס נכסים (Asset Ratio).

    Column R, rows 17–19 (מים, ביוב, סה"כ). Value must exist (not None, empty, or 0).
    """
    ROW_LABELS = ("מים", "ביוב", 'סה"כ')
    results: List[CheckResult] = []

    for i, excel_row in enumerate((17, 18, 19)):
        label = ROW_LABELS[i]
        df_idx = excel_row_to_df_index(excel_row, cfg)
        raw = get_cell(plan_df, df_idx, cfg.value_col_r_idx)
        has_value = raw is not None and not (isinstance(raw, float) and pd.isna(raw))
        if has_value and isinstance(raw, (int, float)):
            try:
                has_value = float(raw) != 0
            except (TypeError, ValueError):
                pass
        elif has_value:
            has_value = str(raw).strip() not in ("", "0")

        status = Status.PASS_ if has_value else Status.FAIL
        if status == Status.FAIL:
            message = f"חסר ערך או ערך 0 ({label}) R{excel_row}"
        else:
            message = f"ערך קיים ({label}) R{excel_row}"
        actual_val = raw
        if has_value and raw is not None:
            try:
                actual_val = round(float(raw), 2)
            except (TypeError, ValueError):
                pass
        results.append(
            CheckResult(
                rule_id=f"R_2_{label.replace(chr(34), '')}",
                rule_name="גריעת נכסים + פרטים",
                severity=Severity.CRITICAL,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name="R",
                key_context=f"plan_cell=R{excel_row}; value must exist and not 0",
                actual_value=actual_val,
                expected_value="ערך קיים ולא 0",
                status=status,
                message=message,
                excel_cells=[_summary_sheet_cell(cfg, excel_row, "R")],
            )
        )

    return results


def check_003_defined_value_percent(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    Rule R_3: לציין את הערך שהוגדר.

    Column R, rows 20–22 (מים, ביוב, סה"כ). Value must exist and be between 0 and 100 (%).
    """
    ROW_LABELS = ("מים", "ביוב", 'סה"כ')
    results: List[CheckResult] = []

    for i, excel_row in enumerate((20, 21, 22)):
        label = ROW_LABELS[i]
        df_idx = excel_row_to_df_index(excel_row, cfg)
        raw = get_cell(plan_df, df_idx, cfg.value_col_r_idx)
        try:
            val = float(raw) if raw is not None else None
        except (TypeError, ValueError):
            val = None

        # Cell stores ratio (e.g. 0.49 = 49%); do not divide — use as percent: 0.49 → 49%
        if val is not None and 0 < val <= 1:
            pct = val * 100
        elif val is not None:
            pct = val
        else:
            pct = None

        if pct is None or (isinstance(raw, float) and pd.isna(raw)):
            status = Status.FAIL
            message = f"חסר ערך ({label}) R{excel_row}"
            actual_value = raw
            expected_value = "ערך בין 0 ל-100%"
        elif not (0 <= pct <= 100):
            status = Status.FAIL
            message = f"ערך מחוץ לטווח 0–100% ({label}) R{excel_row}: {round(pct, 2)}%"
            actual_value = f"{round(pct, 2)}%"
            expected_value = "ערך בין 0 ל-100%"
        else:
            status = Status.PASS_
            message = f"ערך בטווח ({label}) R{excel_row}: {round(pct, 2)}%"
            actual_value = f"{round(pct, 2)}%"
            expected_value = "ערך בין 0 ל-100%"

        results.append(
            CheckResult(
                rule_id=f"R_3_{label.replace(chr(34), '')}",
                rule_name="סה\"כ  נתוני תוכנית השקעה -יחס נכסים",
                severity=Severity.CRITICAL,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name="R",
                key_context=f"plan_cell=R{excel_row}; value 0-100%",
                actual_value=actual_value,
                expected_value=expected_value,
                status=status,
                message=message,
                excel_cells=[_summary_sheet_cell(cfg, excel_row, "R")],
            )
        )

    return results


def check_004_total_program_values(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    Rule R_4: מינימום נדרש + שיקום ושדרוג (מימון עצמי) + ratio.

    Sub-check 1: Column R rows 25,26,27 (מינימום נדרש) — must have values. Labels: מים, ביוב, סה"כ.
    Sub-check 1b: Column R rows 28,29,30 (שיקום ושדרוג מימון עצמי) — must have values.
    Sub-check 2: Column S rows 28,29,30 — value is ratio (e.g. 2.44 = 244.2%). Display 1 decimal, must be > 100%.
    """
    ROW_LABELS = ("מים", "ביוב", 'סה"כ')
    results: List[CheckResult] = []

    # Sub-check 1: מינימום נדרש — Column R, rows 25, 26, 27
    for i, excel_row in enumerate((25, 26, 27)):
        label = ROW_LABELS[i]
        df_idx = excel_row_to_df_index(excel_row, cfg)
        val = get_cell(plan_df, df_idx, cfg.value_col_r_idx)
        has_value = val is not None and not (isinstance(val, float) and pd.isna(val))
        if has_value and isinstance(val, (int, float)):
            try:
                has_value = float(val) != 0
            except (TypeError, ValueError):
                pass
        elif has_value:
            has_value = str(val).strip() not in ("", "0")

        status = Status.PASS_ if has_value else Status.FAIL
        message = f"מינימום נדרש ({label}) R{excel_row}: {'ערך קיים' if has_value else 'חסר ערך'}"
        try:
            actual_val = round(float(val), 2) if val is not None else None
        except (TypeError, ValueError):
            actual_val = val
        results.append(
            CheckResult(
                rule_id=f"R_4_מינימום_{label.replace(chr(34), '')}",
                rule_name='מינימום נדרש / שיקום ושדרוג',
                severity=Severity.CRITICAL,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name="R",
                key_context=f"plan_cell=R{excel_row}; מינימום נדרש",
                actual_value=actual_val,
                expected_value="ערך קיים",
                status=status,
                message=message,
                excel_cells=[_summary_sheet_cell(cfg, excel_row, "R")],
            )
        )

    # Sub-check 1b: שיקום ושדרוג (מימון עצמי) — Column R, rows 28, 29, 30
    for i, excel_row in enumerate((28, 29, 30)):
        label = ROW_LABELS[i]
        df_idx = excel_row_to_df_index(excel_row, cfg)
        val = get_cell(plan_df, df_idx, cfg.value_col_r_idx)
        has_value = val is not None and not (isinstance(val, float) and pd.isna(val))
        if has_value and isinstance(val, (int, float)):
            try:
                has_value = float(val) != 0
            except (TypeError, ValueError):
                pass
        elif has_value:
            has_value = str(val).strip() not in ("", "0")

        status = Status.PASS_ if has_value else Status.FAIL
        message = f"שיקום ושדרוג מימון עצמי ({label}) R{excel_row}: {'ערך קיים' if has_value else 'חסר ערך'}"
        try:
            actual_val = round(float(val), 2) if val is not None else None
        except (TypeError, ValueError):
            actual_val = val
        results.append(
            CheckResult(
                rule_id=f"R_4_שיקום_{label.replace(chr(34), '')}",
                rule_name='מינימום נדרש / שיקום ושדרוג',
                severity=Severity.CRITICAL,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name="R",
                key_context=f"plan_cell=R{excel_row}; שיקום ושדרוג מימון עצמי",
                actual_value=actual_val,
                expected_value="ערך קיים",
                status=status,
                message=message,
                excel_cells=[_summary_sheet_cell(cfg, excel_row, "R")],
            )
        )

    # Sub-check 2: Ratio in Column S rows 28,29,30 — stored as ratio (e.g. 2.44 = 244%); must be > 100%
    for i, excel_row in enumerate((28, 29, 30)):
        label = ROW_LABELS[i]
        df_idx = excel_row_to_df_index(excel_row, cfg)
        raw = get_cell(plan_df, df_idx, cfg.value_col_s_idx)
        try:
            ratio_val = float(raw) if raw is not None else None
        except (TypeError, ValueError):
            ratio_val = None

        if ratio_val is None:
            status = Status.FAIL
            message = f"יחס מינימום נדרש לשיקום ושדרוג ({label}) S{excel_row}: חסר או לא חוקי. נדרש מעל 100%"
            actual_value = raw
            expected_value = "> 100%"
        else:
            pct_val = ratio_val * 100  # convert ratio to percent (2.44 -> 244.2)
            pct_display = round(pct_val, 1)
            ok = pct_val > 100
            status = Status.PASS_ if ok else Status.FAIL
            actual_value = f"{pct_display}%"
            expected_value = "> 100%"
            if ok:
                message = f"יחס מינימום נדרש לשיקום ושדרוג ({label}): {pct_display}%"
            else:
                message = f"יחס מינימום נדרש לשיקום ושדרוג ({label}) הוא {pct_display}%. מתחת ל-100% דורש בירור"

        results.append(
            CheckResult(
                rule_id=f"R_4_יחס_{label.replace(chr(34), '')}",
                rule_name='יחס נכסים',
                severity=Severity.CRITICAL,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name="S",
                key_context=f"plan_cell=S{excel_row}; ratio מינימום נדרש לשיקום ושדרוג",
                actual_value=actual_value,
                expected_value=expected_value,
                status=status,
                message=message,
                excel_cells=[_summary_sheet_cell(cfg, excel_row, "S")],
            )
        )

    return results




def check_005_total_planned_investments_cross_row(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    R_5: דיווח סה"כ השקעות מתוכננות לביצוע.

    Cross-row check for columns C–Q (indices 2–16): Row 4 = city name, Row 39 = investment value.
    If Row 4 has a city name, Row 39 MUST have a valid non-zero numerical value.
    """
    ROW_CITY = 4
    ROW_INVESTMENT = 39
    COL_START, COL_END = 2, 17  # C=2 through Q=16

    df_idx_4 = excel_row_to_df_index(ROW_CITY, cfg)
    df_idx_39 = excel_row_to_df_index(ROW_INVESTMENT, cfg)
    failed_cities: List[str] = []

    for col_idx in range(COL_START, COL_END):
        city_name = get_cell(plan_df, df_idx_4, col_idx)
        value = get_cell(plan_df, df_idx_39, col_idx)

        has_city = (
            city_name is not None
            and not (isinstance(city_name, float) and pd.isna(city_name))
            and str(city_name).strip() != ""
        )
        if not has_city:
            continue

        city_str = str(city_name).strip()
        valid_value = False
        if value is not None and not (isinstance(value, float) and pd.isna(value)) and str(value).strip() != "":
            try:
                n = float(value)
                valid_value = n != 0
            except (TypeError, ValueError):
                pass
        if not valid_value:
            failed_cities.append(city_str)

    status = Status.FAIL if failed_cities else Status.PASS_
    if failed_cities:
        parts = [f'לישוב {city} אין דיווח סה"כ מתוכננות לפרויקטים מתוכננים' for city in failed_cities]
        message = "נכשל - במידה והערך לא עומד בתנאי יש להציג הודעה: " + "; ".join(parts)
        actual_value = failed_cities
        expected_value = "ערך מספרי לא ריק ולא 0 בשורה 39 לכל יישוב"
    else:
        message = "תקין - לכל היישובים המדווחים קיים ערך בשורה 39."
        actual_value = "כל היישובים תקינים"
        expected_value = "ערך מספרי לא ריק ולא 0 בשורה 39 לכל יישוב"

    col_letters = "C–Q"
    excel_cells = [_summary_sheet_cell(cfg, ROW_INVESTMENT, _col_idx_to_letter(c)) for c in range(COL_START, COL_END)]
    results: List[CheckResult] = [
        CheckResult(
            rule_id="R_5",
            rule_name='דיווח סה"כ השקעות מתוכננות לביצוע',
            severity=Severity.CRITICAL,
            sheet_name=cfg.sheet_name,
            row_index=df_idx_39,
            column_name=col_letters,
            key_context=f"row_4=יישוב; row_39=ערך השקעה; עמודות {col_letters}",
            actual_value=actual_value,
            expected_value=expected_value,
            status=status,
            message=message,
            excel_cells=excel_cells,
        )
    ]
    return results


def check_005_min_required_program(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    Minimum required program values in column R. FAIL if missing.
    """
    results: List[CheckResult] = []

    for system, excel_row in cfg.min_required_program_rows_excel.items():
        df_idx = excel_row_to_df_index(excel_row, cfg)
        val = get_cell(plan_df, df_idx, cfg.value_col_r_idx)

        results.append(
            CheckResult(
                rule_id=f"R_005_{system}",
                rule_name="מינימום נדרש לתכנית השקעה",
                severity=Severity.CRITICAL,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name="R",
                key_context=f"plan_cell=R{excel_row}",
                actual_value=val,
                expected_value="reported value",
                status=Status.PASS_ if pd.notna(val) else Status.FAIL,
                message=f"Value from R{excel_row} = {val}",
                excel_cells=[_summary_sheet_cell(cfg, excel_row, "R")],
            )
        )

    return results


def _sync_investment_funding_scan(plan_df: pd.DataFrame, cfg: PlanConfig) -> tuple[list[str], list[str]]:
    """
    Scan columns C–Q (2–16): row 4 = city, row 39 = investment, row 50 = budget.
    Returns (missing_sources_failures, deficit_failures) where each is a list of error messages.
    """
    ROW_CITY, ROW_INVESTMENT, ROW_BUDGET = 4, 39, 50
    COL_START, COL_END = 2, 17

    df_idx_4 = excel_row_to_df_index(ROW_CITY, cfg)
    df_idx_39 = excel_row_to_df_index(ROW_INVESTMENT, cfg)
    df_idx_50 = excel_row_to_df_index(ROW_BUDGET, cfg)

    missing_sources: list[str] = []
    deficit: list[str] = []

    for col_idx in range(COL_START, COL_END):
        city_name = get_cell(plan_df, df_idx_4, col_idx)
        has_city = (
            city_name is not None
            and not (isinstance(city_name, float) and pd.isna(city_name))
            and str(city_name).strip() != ""
        )
        if not has_city:
            continue

        city_str = str(city_name).strip()
        inv_raw = get_cell(plan_df, df_idx_39, col_idx)
        bud_raw = get_cell(plan_df, df_idx_50, col_idx)

        try:
            inv_val = float(inv_raw) if inv_raw is not None and not (isinstance(inv_raw, float) and pd.isna(inv_raw)) and str(inv_raw).strip() != "" else None
        except (TypeError, ValueError):
            inv_val = None
        try:
            bud_val = float(bud_raw) if bud_raw is not None and not (isinstance(bud_raw, float) and pd.isna(bud_raw)) and str(bud_raw).strip() != "" else None
        except (TypeError, ValueError):
            bud_val = None

        # R_6: Row 50 empty, null, or 0
        if bud_val is None or bud_val == 0:
            missing_sources.append(f"לא קיים ערך של מקורות תקציב עבור רשות מקומית {city_str}")

        # R_7: Both exist, investment > budget
        if inv_val is not None and bud_val is not None and bud_val != 0:
            if inv_val - bud_val > 0:
                deficit.append(f"אין מספיק מקורות מימון - רשות מקומית {city_str}")

    return missing_sources, deficit


def check_006_sync_budget_sources_missing(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    R_6: בדיקת סנכרון השקעות ומקורות מימון.
    Sub-Check A: For each city in row 4 (columns C–Q), row 50 (Total Budget) must not be empty/null/0.
    """
    missing_sources, _ = _sync_investment_funding_scan(plan_df, cfg)
    ROW_BUDGET = 50
    COL_START, COL_END = 2, 17

    status = Status.FAIL if missing_sources else Status.PASS_
    message = "; ".join(missing_sources) if missing_sources else "תקין - לכל היישובים קיים ערך מקורות תקציב בשורה 50."
    df_idx_50 = excel_row_to_df_index(ROW_BUDGET, cfg)
    excel_cells = [_summary_sheet_cell(cfg, ROW_BUDGET, _col_idx_to_letter(c)) for c in range(COL_START, COL_END)]

    return [
        CheckResult(
            rule_id="R_6",
            rule_name="בדיקת סנכרון השקעות ומקורות מימון",
            severity=Severity.CRITICAL,
            sheet_name=cfg.sheet_name,
            row_index=df_idx_50,
            column_name="C–Q",
            key_context="row_4=יישוב; row_50=מקורות תקציב; עמודות C–Q",
            actual_value=missing_sources if missing_sources else "כל היישובים עם ערך מקורות תקציב",
            expected_value="ערך מקורות תקציב לא ריק ולא 0 לכל יישוב",
            status=status,
            message=message,
            excel_cells=excel_cells,
        )
    ]


def check_007_sync_budget_deficit(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    R_7: בדיקת סנכרון השקעות ומקורות מימון.
    Sub-Check B: For each city, if both row 39 (investment) and row 50 (budget) exist, investment must not exceed budget.
    """
    _, deficit = _sync_investment_funding_scan(plan_df, cfg)
    ROW_INVESTMENT, ROW_BUDGET = 39, 50
    COL_START, COL_END = 2, 17

    status = Status.FAIL if deficit else Status.PASS_
    message = "; ".join(deficit) if deficit else "תקין - מקורות המימון מכסים את ההשקעות המתוכננות לכל יישוב."
    df_idx_39 = excel_row_to_df_index(ROW_INVESTMENT, cfg)
    excel_cells = [_summary_sheet_cell(cfg, ROW_INVESTMENT, _col_idx_to_letter(c)) for c in range(COL_START, COL_END)]

    return [
        CheckResult(
            rule_id="R_7",
            rule_name="בדיקת סנכרון השקעות ומקורות מימון",
            severity=Severity.CRITICAL,
            sheet_name=cfg.sheet_name,
            row_index=df_idx_39,
            column_name="C–Q",
            key_context="row_39=השקעה מתוכננת; row_50=מקורות תקציב; עמודות C–Q",
            actual_value=deficit if deficit else "מקורות מימון מכסים את ההשקעות",
            expected_value="מקורות מימון >= השקעה מתוכננת לכל יישוב",
            status=status,
            message=message,
            excel_cells=excel_cells,
        )
    ]


def check_006_rehab_upgrade_min_required(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    Minimum required rehab/upgrade values in column S. FAIL if missing.
    """
    results: List[CheckResult] = []

    for system, excel_row in cfg.rehab_upgrade_min_rows_excel.items():
        df_idx = excel_row_to_df_index(excel_row, cfg)
        val = get_cell(plan_df, df_idx, cfg.value_col_s_idx)

        results.append(
            CheckResult(
                rule_id=f"R_006_{system}",
                rule_name="מינימום נדרש לשיקום/שדרוג",
                severity=Severity.INFO,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name="S",
                key_context=f"plan_cell=S{excel_row}",
                actual_value=val,
                expected_value="reported value",
                status=Status.PASS_ if pd.notna(val) else Status.FAIL,
                message=f"Value from S{excel_row} = {val}",
                excel_cells=[_summary_sheet_cell(cfg, excel_row, "S")],
            )
        )

    return results


def check_007_total_planned_investments_by_city(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    City columns: check row planned_investments_row_excel has values per city.
    FAIL per city if blank.
    """
    city_cols = detect_city_cols(plan_df)
    if not city_cols:
        return fail_no_cities("R_007", 'סה"כ השקעות מתוכננות לביצוע', cfg)

    excel_row = cfg.planned_investments_row_excel
    df_idx = excel_row_to_df_index(excel_row, cfg)

    results: List[CheckResult] = []
    for city in city_cols:
        val = plan_df.at[df_idx, city] if df_idx in plan_df.index else None
        col_idx = plan_df.columns.get_loc(city)
        col_letter = _col_idx_to_letter(col_idx)
        results.append(
            CheckResult(
                rule_id=f"R_007_{city}",
                rule_name='סה"כ השקעות מתוכננות לביצוע',
                severity=Severity.CRITICAL,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name=str(city),
                key_context=f"row={excel_row}; city={city}",
                actual_value=val,
                expected_value="reported value",
                status=Status.PASS_ if pd.notna(val) else Status.FAIL,
                message=f"Value from row {excel_row} for '{city}' = {val}",
                excel_cells=[_summary_sheet_cell(cfg, excel_row, col_letter)],
            )
        )
    return results


def check_008_pipe_lengths_water(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    R_8: דיווח אורכי צנרת מים.
    For each column C–Q: row 4 = city. Skip 'כפר סבא'. For other cities, at least one of row 56 or 57
    must have a valid non-zero numerical value (pipe length). FAIL if both 56 and 57 are empty/zero.
    """
    ROW_CITY, ROW_56, ROW_57 = 4, 56, 57
    COL_START, COL_END = 2, 17
    SKIP_CITY = "כפר סבא"

    df_idx_4 = excel_row_to_df_index(ROW_CITY, cfg)
    df_idx_56 = excel_row_to_df_index(ROW_56, cfg)
    df_idx_57 = excel_row_to_df_index(ROW_57, cfg)

    failed_cities: List[str] = []

    for col_idx in range(COL_START, COL_END):
        city_name = get_cell(plan_df, df_idx_4, col_idx)
        has_city = (
            city_name is not None
            and not (isinstance(city_name, float) and pd.isna(city_name))
            and str(city_name).strip() != ""
        )
        if not has_city:
            continue

        city_str = str(city_name).strip()
        if city_str == SKIP_CITY:
            continue

        v56 = get_cell(plan_df, df_idx_56, col_idx)
        v57 = get_cell(plan_df, df_idx_57, col_idx)

        def valid_num(val: Any) -> bool:
            if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
                return False
            try:
                return float(val) != 0
            except (TypeError, ValueError):
                return False

        if not valid_num(v56) and not valid_num(v57):
            failed_cities.append(city_str)

    status = Status.FAIL if failed_cities else Status.PASS_
    if failed_cities:
        parts = [f"לישוב {c} חסר דיווח אורכי צנרת מים (שורה 56 או 57)" for c in failed_cities]
        message = "נכשל - " + "; ".join(parts)
        actual_value = failed_cities
        expected_value = "ערך לא ריק ולא 0 בשורה 56 או 57 לכל יישוב (מלבד כפר סבא)"
    else:
        message = "תקין - לכל היישובים הנדרשים קיים דיווח אורכי צנרת מים (שורה 56 או 57)."
        actual_value = "כל היישובים דיווחו"
        expected_value = "ערך לא ריק ולא 0 בשורה 56 או 57 לכל יישוב (מלבד כפר סבא)"

    excel_cells = []
    for r in (ROW_56, ROW_57):
        for c in range(COL_START, COL_END):
            excel_cells.append(_summary_sheet_cell(cfg, r, _col_idx_to_letter(c)))

    return [
        CheckResult(
            rule_id="R_8",
            rule_name="דיווח אורכי צנרת מים",
            severity=Severity.CRITICAL,
            sheet_name=cfg.sheet_name,
            row_index=df_idx_56,
            column_name="C–Q",
            key_context="row_4=יישוב; row_56,57=אורכי צנרת מים; עמודות C–Q; כפר סבא מדולג",
            actual_value=actual_value,
            expected_value=expected_value,
            status=status,
            message=message,
            excel_cells=excel_cells,
        )
    ]


def check_008_funding_total_and_exists_by_city(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    City columns: check row funding_total_row_excel exists/non-empty per city.
    FAIL per city if empty.
    """
    city_cols = detect_city_cols(plan_df)
    if not city_cols:
        return fail_no_cities("R_008", 'מקורות מימון - סה"כ מקורות תקציב', cfg)

    excel_row = cfg.funding_total_row_excel
    df_idx = excel_row_to_df_index(excel_row, cfg)

    results: List[CheckResult] = []
    for city in city_cols:
        val = plan_df.at[df_idx, city] if df_idx in plan_df.index else None
        exists = is_non_empty(val)
        col_idx = plan_df.columns.get_loc(city)
        col_letter = _col_idx_to_letter(col_idx)
        results.append(
            CheckResult(
                rule_id=f"R_008_{city}",
                rule_name='מקורות מימון - סה"כ מקורות תקציב (כולל בדיקת קיום)',
                severity=Severity.CRITICAL,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name=str(city),
                key_context=f"row={excel_row}; city={city}",
                actual_value=val,
                expected_value="non-empty",
                status=Status.PASS_ if exists else Status.FAIL,
                message=f"Funding total for '{city}' from row {excel_row} = {val}",
                excel_cells=[_summary_sheet_cell(cfg, excel_row, col_letter)],
            )
        )
    return results


def check_009_pipe_lengths_sewer(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    R_9: דיווח אורכי צנרת ביוב.
    For each column C–Q: row 4 = city. Skip 'כפר סבא'. For other cities, row 58 must have
    a valid non-zero numerical value (sewage pipe length). FAIL if row 58 is empty or zero.
    """
    ROW_CITY, ROW_58 = 4, 58
    COL_START, COL_END = 2, 17
    SKIP_CITY = "כפר סבא"

    df_idx_4 = excel_row_to_df_index(ROW_CITY, cfg)
    df_idx_58 = excel_row_to_df_index(ROW_58, cfg)

    failed_cities: List[str] = []

    for col_idx in range(COL_START, COL_END):
        city_name = get_cell(plan_df, df_idx_4, col_idx)
        has_city = (
            city_name is not None
            and not (isinstance(city_name, float) and pd.isna(city_name))
            and str(city_name).strip() != ""
        )
        if not has_city:
            continue

        city_str = str(city_name).strip()
        if city_str == SKIP_CITY:
            continue

        v58 = get_cell(plan_df, df_idx_58, col_idx)
        valid = False
        if v58 is not None and not (isinstance(v58, float) and pd.isna(v58)) and str(v58).strip() != "":
            try:
                valid = float(v58) != 0
            except (TypeError, ValueError):
                pass
        if not valid:
            failed_cities.append(city_str)

    status = Status.FAIL if failed_cities else Status.PASS_
    if failed_cities:
        parts = [f"לישוב {c} חסר דיווח אורכי צנרת ביוב (שורה 58)" for c in failed_cities]
        message = "נכשל - " + "; ".join(parts)
        actual_value = failed_cities
        expected_value = "ערך לא ריק ולא 0 בשורה 58 לכל יישוב (מלבד כפר סבא)"
    else:
        message = "תקין - לכל היישובים הנדרשים קיים דיווח אורכי צנרת ביוב (שורה 58)."
        actual_value = "כל היישובים דיווחו"
        expected_value = "ערך לא ריק ולא 0 בשורה 58 לכל יישוב (מלבד כפר סבא)"

    excel_cells = [_summary_sheet_cell(cfg, ROW_58, _col_idx_to_letter(c)) for c in range(COL_START, COL_END)]

    return [
        CheckResult(
            rule_id="R_9",
            rule_name="דיווח אורכי צנרת ביוב",
            severity=Severity.CRITICAL,
            sheet_name=cfg.sheet_name,
            row_index=df_idx_58,
            column_name="C–Q",
            key_context="row_4=יישוב; row_58=אורכי צנרת ביוב; עמודות C–Q; כפר סבא מדולג",
            actual_value=actual_value,
            expected_value=expected_value,
            status=status,
            message=message,
            excel_cells=excel_cells,
        )
    ]


def check_010_pipes_any_value(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    R_010: For each city, check that at least ONE of:
      - row 56 (water steel / PE)
      - row 57 (water PVC)
      - row 58 (sewer)
    has a value.
    """
    city_cols = detect_city_cols(plan_df)
    if not city_cols:
        return fail_no_cities("R_010", "דיווח אורכי צנרת (לפחות ערך אחד)", cfg)

    row_ws, row_wp = cfg.water_pipe_rows_excel
    row_sw = cfg.sewer_pipe_row_excel

    idx_ws = excel_row_to_df_index(row_ws, cfg)
    idx_wp = excel_row_to_df_index(row_wp, cfg)
    idx_sw = excel_row_to_df_index(row_sw, cfg)

    results: List[CheckResult] = []

    for city in city_cols:
        v_ws = plan_df.at[idx_ws, city] if idx_ws in plan_df.index else None
        v_wp = plan_df.at[idx_wp, city] if idx_wp in plan_df.index else None
        v_sw = plan_df.at[idx_sw, city] if idx_sw in plan_df.index else None

        ok = is_non_empty(v_ws) or is_non_empty(v_wp) or is_non_empty(v_sw)
        col_idx = plan_df.columns.get_loc(city)
        col_letter = _col_idx_to_letter(col_idx)
        excel_cells_list = [
            _summary_sheet_cell(cfg, row_ws, col_letter),
            _summary_sheet_cell(cfg, row_wp, col_letter),
            _summary_sheet_cell(cfg, row_sw, col_letter),
        ]

        results.append(
            CheckResult(
                rule_id=f"R_010_{city}",
                rule_name="דיווח אורכי צנרת (לפחות ערך אחד מתוך 3 שורות)",
                severity=Severity.WARNING,
                sheet_name=cfg.sheet_name,
                row_index=idx_ws,
                column_name=str(city),
                key_context=f"rows={row_ws},{row_wp},{row_sw}; city={city}",
                actual_value={"water_row56": v_ws, "water_row57": v_wp, "sewer_row58": v_sw},
                expected_value="at least one non-empty among rows 56/57/58",
                status=Status.PASS_ if ok else Status.FAIL,
                message=f"row56={v_ws}, row57={v_wp}, row58={v_sw}",
                excel_cells=excel_cells_list,
            )
        )

    return results


def check_011_pipes_values_by_type(plan_df: pd.DataFrame, cfg: PlanConfig) -> List[CheckResult]:
    """
    R_011: Output pipe lengths by type (3 rows per city):
      - row 56: water steel / PE
      - row 57: water PVC
      - row 58: sewer
    """
    city_cols = detect_city_cols(plan_df)
    if not city_cols:
        return fail_no_cities("R_011", "דיווח אורכי צנרת (פירוט ערכים)", cfg)

    row_ws, row_wp = cfg.water_pipe_rows_excel
    row_sw = cfg.sewer_pipe_row_excel

    idx_ws = excel_row_to_df_index(row_ws, cfg)
    idx_wp = excel_row_to_df_index(row_wp, cfg)
    idx_sw = excel_row_to_df_index(row_sw, cfg)

    results: List[CheckResult] = []

    def emit(city: str, pipe_type: str, excel_row: int, df_idx: int, val: Any, suffix: str) -> None:
        col_idx = plan_df.columns.get_loc(city)
        col_letter = _col_idx_to_letter(col_idx)
        results.append(
            CheckResult(
                rule_id=f"R_011_{suffix}_{city}",
                rule_name=f"דיווח אורכי צנרת - {pipe_type}",
                severity=Severity.INFO,
                sheet_name=cfg.sheet_name,
                row_index=df_idx,
                column_name=str(city),
                key_context=f"row={excel_row}; city={city}; type={pipe_type}",
                actual_value=val,
                expected_value="reported value",
                status=Status.PASS_,  # reporting
                message=f"Value from row {excel_row} for '{city}' ({pipe_type}) = {val}",
                excel_cells=[_summary_sheet_cell(cfg, excel_row, col_letter)],
            )
        )

    for city in city_cols:
        v1 = plan_df.at[idx_ws, city] if idx_ws in plan_df.index else None
        v2 = plan_df.at[idx_wp, city] if idx_wp in plan_df.index else None
        v3 = plan_df.at[idx_sw, city] if idx_sw in plan_df.index else None

        emit(city, "מים פלדה (PE/פלדה)", row_ws, idx_ws, v1, "WATER_STEEL")
        emit(city, "מים PVC", row_wp, idx_wp, v2, "WATER_PVC")
        emit(city, "ביוב", row_sw, idx_sw, v3, "SEWER")

    return results


def check_012_project_fields_not_empty(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    rule12_projectName_Not_NA:
    In sheet "גיליון דיווח" ensure columns:
      - שם פרויקט
      - מיקום פרויקט
      - סיווג פרויקט
    are always filled for real data rows.

    Note: Excel has messy headers ("מיקום פרויקט " with trailing space, "סיווג\nפרויקט").
    We normalize column names to match robustly.
    """

    checked_cols_label = "שם פרויקט / מיקום פרויקט / סיווג פרויקט"

    rule_id = "R_12"
    rule_name = "שם פרויקט / מיקום פרויקט / סיווג פרויקט - חובה (Not NA)"

    def _norm_col(c: str) -> str:
        return re.sub(r"\s+", " ", str(c)).strip()

    # Map normalized -> original
    norm_to_orig = {_norm_col(c): c for c in report_df.columns}

    required_norm = ["שם פרויקט", "מיקום פרויקט", "סיווג פרויקט"]
    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")

    missing_cols = [c for c in required_norm + [id_norm] if c not in norm_to_orig]
    if missing_cols:
        return [
            CheckResult(
                rule_id=rule_id,
                rule_name=rule_name,
                severity=Severity.CRITICAL,
                sheet_name=getattr(cfg, "report_sheet_name", "גיליון דיווח"),
                row_index=None,
                column_name=None,
                key_context="columns_presence",
                actual_value=list(report_df.columns),
                expected_value=f"columns include: {required_norm + [id_norm]}",
                status=Status.FAIL,
                message=f"Missing columns (after normalization): {missing_cols}",
            )
        ]

    id_col = norm_to_orig[id_norm]
    req_cols = [norm_to_orig[c] for c in required_norm]

    # Decide which rows are “real”
    # In your file: non-data rows have '-' in מס' פרויקט
    id_series = report_df[id_col].astype(str).str.strip()
    data_mask = report_df[id_col].notna() & (id_series != "-") & (id_series != "nan") & (id_series != "")

    df_data = report_df.loc[data_mask, [id_col] + req_cols].copy()

    def is_missing(v: Any) -> bool:
        if pd.isna(v):
            return True
        if isinstance(v, str) and v.strip() == "":
            return True
        return False

    miss = df_data[req_cols].map(is_missing)
    failed_mask = miss.any(axis=1)

    # Convert failures into CheckResult rows (one per failing row)
    results: List[CheckResult] = []

    # Excel row number: header_row=6 => header is Excel row 7.
    # DataFrame index 0 corresponds to Excel row 8.
    header_row = getattr(cfg, "report_header_row", 6)
    excel_row_offset = (header_row + 3)  # header_row(6) + 1 (1-based) + 1 (header) + 1 (sub-header) = row 9 for df_idx 0


    from openpyxl.utils import get_column_letter

    # Map original df columns -> Excel column letters (A,B,C,...)
    col_to_excel_letter = {
        col: get_column_letter(i + 1)
        for i, col in enumerate(list(report_df.columns))
    }

    for df_idx, row in df_data.loc[failed_mask].iterrows():
        missing_fields = [ _norm_col(col) for col, flag in miss.loc[df_idx].items() if flag ]
        excel_row = int(df_idx + excel_row_offset)
        excel_cells = []
        for col, flag in miss.loc[df_idx].items():
            if flag:
                # col here is original column name (like "מיקום פרויקט ")
                letter = col_to_excel_letter.get(col)
                if letter:
                    excel_cells.append(f"{getattr(cfg, 'report_sheet_name', 'גיליון דיווח')}!{letter}{excel_row}")
        
        results.append(
            CheckResult(
                rule_id=rule_id,
                rule_name=rule_name,
                severity=Severity.CRITICAL,
                sheet_name=getattr(cfg, "report_sheet_name", "גיליון דיווח"),
                row_index=df_idx,
                column_name=" / ".join(required_norm),
                key_context=f"{id_norm}={row[id_col]} | excel_row={excel_row} | checked_cols={checked_cols_label}",
                actual_value={_norm_col(c): row[c] for c in req_cols},
                expected_value="non-empty values",
                status=Status.FAIL,
                message=f"Missing fields: {missing_fields} (Excel row {excel_row})",
                excel_cells=excel_cells,   # ✅ NEW
            )
        )

    # If no failures, emit a single PASS result (so it shows nicely in summary)
    if not results:
        results.append(
            CheckResult(
                rule_id=rule_id,
                rule_name=rule_name,
                severity=Severity.INFO,
                sheet_name=getattr(cfg, "report_sheet_name", "גיליון דיווח"),
                row_index=None,
                column_name=" / ".join(required_norm),
                key_context=f"rows_checked={len(df_data)} | checked_cols={checked_cols_label}",
                actual_value=len(df_data),
                expected_value="all required fields filled",
                status=Status.PASS_,
                message="All required project fields are filled.",
            )
        )

    return results

# =============================================================================
# Rule 14: Funding Classification Validation (LLM/NLP)
# =============================================================================

# Fuzzy threshold constant
R14_FUZZY_THRESHOLD = 0.50  # Confidence >= 50% -> accept fuzzy decision


def canonicalize_label(s: str) -> str:
    """
    Canonicalize a funding label to one of the 5 allowed canonical forms.

    Canonical labels:
    - "שיקום/שדרוג"
    - "פיתוח"
    - "תשתית ביוב אזורית"
    - "קידוחים"
    - "תחזוקה / שוטף"
    """
    if not s:
        return ""

    # Normalize whitespace (including NBSP)
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()

    # --- תשתית ביוב אזורית variations (singular canonical) ---
    if "תשתיות ביוב אזוריות" in s:
        s = s.replace("תשתיות ביוב אזוריות", "תשתית ביוב אזורית")
    if "תשתיות ביוב אזורית" in s:
        s = s.replace("תשתיות ביוב אזורית", "תשתית ביוב אזורית")
    if "תשתית ביוב אזוריות" in s:
        s = s.replace("תשתית ביוב אזוריות", "תשתית ביוב אזורית")
    if "תשתית ביוב אזורי" in s and "תשתית ביוב אזורית" not in s:
        s = s.replace("תשתית ביוב אזורי", "תשתית ביוב אזורית")

    # --- קידוחים variations ---
    # Only replace standalone "קידוח" when it's likely to be the label
    if s == "קידוח":
        s = "קידוחים"

    # --- שיקום/שדרוג variations ---
    s = s.replace("שיקום ושדרוג", "שיקום/שדרוג")
    s = s.replace("שיקום ושידרוג", "שיקום/שדרוג")
    s = s.replace("שיקום / שדרוג", "שיקום/שדרוג")
    s = s.replace("שיקום/ שדרוג", "שיקום/שדרוג")
    s = s.replace("שיקום /שדרוג", "שיקום/שדרוג")

    # --- תחזוקה / שוטף variations (canonical has spaces around slash) ---
    # Match various spacing patterns around the slash
    s = re.sub(r"תחזוקה\s*/\s*שוטף", "תחזוקה / שוטף", s)
    s = s.replace("תחזוקה ושוטף", "תחזוקה / שוטף")

    return s


def _token_set(s: str) -> set[str]:
    """Split string into word tokens."""
    return set(s.split())


def _best_subject_match(project_name: str) -> tuple[str, float]:
    """
    Find the best matching subject from SUBJECT_TO_ALLOWED using fuzzy matching.

    Returns: (best_subject, score) where score is in [0..1]
    """
    pn = normalize_text(project_name)
    pt = _token_set(pn)

    best_subj = ""
    best_score = -1.0

    for subj in SUBJECT_TO_ALLOWED.keys():
        sn = normalize_text(subj)
        st = _token_set(sn)

        # Jaccard similarity on word tokens
        jacc = (len(pt & st) / len(pt | st)) if (pt or st) else 0.0
        # Sequence similarity
        seq = SequenceMatcher(None, pn, sn).ratio()
        # Combined score
        score = 0.55 * seq + 0.45 * jacc

        if score > best_score:
            best_score = score
            best_subj = subj

    return best_subj, float(best_score)


def check_014_llm_project_funding_classification(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
    utility_name: Optional[str] = None,
) -> List[CheckResult]:
    """
    R14 - Funding Classification Validation (LLM/NLP)

    Validates that the reported funding classification (סיווג פרויקט) matches
    the predicted classification based on project name analysis.

    Pipeline order:
    1. Keyword stage (deterministic, conf=1.0)
    2. Fuzzy stage (synonym matching + fuzzy subject matching)
    3. LLM stage (fallback when keyword + fuzzy fail)
    """
    import os

    # TEMP DEBUG: Show config values at function entry
    print(f"[R_14 DEBUG] check_014 called with cfg.llm_enabled={getattr(cfg, 'llm_enabled', 'NOT SET')}, cfg.llm_provider={getattr(cfg, 'llm_provider', 'NOT SET')}")

    rule_id = "R_14"
    rule_name = "R14 - אימות מקור מימון לפי שם פרויקט (LLM/NLP)"
    sheet = getattr(cfg, "report_sheet_name", "גיליון דיווח")

    # ---------------------------
    # Helper functions
    # ---------------------------
    def _norm_col(x: object) -> str:
        return re.sub(r"\s+", " ", str(x)).strip()

    def _is_empty(v: object) -> bool:
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        if isinstance(v, str) and v.strip() == "":
            return True
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none"}

    def _is_generic_project_name(name: str) -> bool:
        """
        Return True if the name is short (<= 2 words after cleaning) AND lacks
        an explicit action verb. Used to block over-confident LLM classification
        of generic names (structural guard).
        """
        if not name or not isinstance(name, str):
            return False
        cleaned = re.sub(r"\s+", " ", str(name).strip()).strip()
        words = [w for w in cleaned.split() if w]
        if len(words) > 2:
            return False
        # Action verbs (safe): presence of any → not generic
        action_verbs = [
            "החלפה", "החלפת", "שדרוג", "חיזוק", "שיקום", "הקמה", "ביצוע", "הנחה", "הנחת",
        ]
        name_norm = normalize_text(cleaned)
        for verb in action_verbs:
            if verb in name_norm:
                return False
        return True

    def _confidence_bucket(c: float) -> str:
        return "ביטחון גבוה" if c >= R14_FUZZY_THRESHOLD else "ביטחון נמוך"

    def _final_line(method: str, label: str, conf: float) -> str:
        """Build the method detail line (no subject)."""
        if method == "keyword":
            return f"סיווג סופי על ידי keyword: {label}"
        if method == "fuzzy":
            return f"סיווג סופי על ידי fuzzy ({_confidence_bucket(conf)}): {label} (conf={conf:.2f})"
        if method == "llm":
            return f"סיווג סופי על ידי LLM: {label} (conf={conf:.2f})"
        if method == "prior":
            return f"סיווג סופי על ידי prior (ביטחון נמוך): {label} (conf={conf:.2f})"
        if method == "structural_guard":
            return "סיווג סופי: חסום על ידי structural guard (שם כללי)"
        # no-decision case
        return f"סיווג סופי על ידי LLM: לא זמין (conf=0.0)"

    def _predict_by_keyword(pn: str) -> tuple[str | None, float]:
        """
        Keyword-based prediction (Stage A).
        Returns: (label or None, confidence)

        IMPORTANT: Only return high-precision matches here.
        Generic terms like "חדש", "שדרוג" are handled in fuzzy stage as tie-breakers.
        """
        # === קידוחים/באר - highest priority, unambiguous ===
        if ("באר" in pn) or ("קידוח" in pn) or ("קידוחים" in pn) or ("רדיוס מגן" in pn):
            return "קידוחים", 1.0

        # === מט"ש / טיהור שפכים -> תשתית ביוב אזורית ===
        if ('מט"ש' in pn) or ("מטש" in pn) or ("מתקן טיהור" in pn) or ("טיהור שפכים" in pn):
            return "תשתית ביוב אזורית", 1.0

        # === מאסף אזורי / ראשי / חוצה -> תשתית ביוב אזורית ===
        if "מאסף" in pn and ("אזורי" in pn or "ראשי" in pn or "חוצה" in pn):
            return "תשתית ביוב אזורית", 1.0

        # === קו הולכת שפכים/ביוב אזורי -> תשתית ביוב אזורית ===
        if ("הולכת שפכים" in pn) or ("הולכת ביוב" in pn and "אזורי" in pn):
            return "תשתית ביוב אזורית", 1.0

        # === תחזוקה/שוטף - high precision maintenance terms ===
        maintenance_terms = (
            "צילום" in pn or "שטיפה" in pn or
            "סקר " in pn or pn.startswith("סקר") or  # סקר with space or at start (avoid partial)
            "מדידות" in pn or
            'קר"מ' in pn or "קרמ" in pn or "קריאה מרחוק" in pn or
            "GIS" in pn or "gis" in pn.lower() or
            "ניתור" in pn or "ניטור" in pn or
            "מפוחים" in pn or "מפוח" in pn or
            "מדי מים" in pn or "מוני מים" in pn or "החלפת מדים" in pn or
            "סייבר" in pn or "cyber" in pn.lower() or
            "תכנית אב" in pn or "תכניות אב" in pn or
            "הגנה קתודית" in pn or
            "ציוד חירום" in pn or "מחסן חירום" in pn or
            "לכידת אבנית" in pn or
            "הפרדת ביוב" in pn or
            "סריקות הנדסיות" in pn
        )
        if maintenance_terms:
            return "תחזוקה / שוטף", 1.0

        # === פיתוח - strong development zone indicators ===
        # These are almost always new development (פיתוח)
        pituach_strong = (
            ('אז"ת' in pn) or ("אזת" in pn) or
            ("אזור תעשי" in pn) or ("א.ת" in pn) or ("א.ת." in pn) or
            ("היי-טק" in pn) or ("הייטק" in pn) or
            ("מתחם חדש" in pn) or ("שכונה חדשה" in pn) or
            ("ותמ\"ל" in pn) or ('ותמ"ל' in pn) or ("ותמל" in pn) or
            ("פינוי בינוי" in pn) or ("תמא 38" in pn) or ("תמ\"א 38" in pn) or
            # הקמת patterns - "הקמת" + infrastructure = new development
            (pn.startswith("הקמת ") or "הקמת מערכ" in pn or "הקמת תחנ" in pn or
             "הקמת קו" in pn or "הקמת בריכ" in pn or "הקמת מאגר" in pn)
        )
        if pituach_strong:
            return "פיתוח", 0.90

        # === שיקום/שדרוג - strong rehabilitation indicators ===
        # Only return when very clear (action + infra combo)
        shikum_strong = (
            ("שדרוג מאסף" in pn) or ("שדרוג מאספים" in pn) or
            ("שיקום מאסף" in pn) or ("שיקום מאספים" in pn) or
            ("החלפת משאב" in pn) or ("החלפת משאבות" in pn) or
            ("החלפת לוח חשמל" in pn) or
            ("שיקום בריכ" in pn) or ("שיקום מגדל" in pn) or
            ("איטום בריכ" in pn) or ("ציפוי בריכ" in pn) or
            ("שיקום תחנ" in pn) or ("שדרוג תחנ" in pn) or
            ("שיקום מערכ" in pn) or ("שדרוג מערכ" in pn) or
            ("החלפת צנרת" in pn) or ("שיקום צנרת" in pn) or
            ("שיקום קו" in pn) or ("שדרוג קו" in pn)
        )
        if shikum_strong:
            return "שיקום/שדרוג", 0.90

        # === DO NOT return generic שיקום/שדרוג or פיתוח here ===
        # Those are handled by fuzzy stage with subject-aware tie-breaking
        return None, 0.0

    def _subject_from_synonyms_normalized(pn_normalized: str) -> str | None:
        """
        Deterministic synonym matching using normalized text.
        Both the project name and synonym keys are compared in normalized form.
        """
        for key, subject in SUBJECT_SYNONYMS.items():
            key_norm = normalize_text(key)
            if key_norm and (key_norm in pn_normalized):
                return subject
        return None

    def _apply_tiebreaker(pn: str, allowed_set: set[str]) -> str | None:
        """
        Apply keyword-based tie-breaker when allowed_set has exactly 2 options.
        Returns label if tie-breaker succeeds, None otherwise.
        """
        if allowed_set != {"שיקום/שדרוג", "פיתוח"}:
            return None  # Only apply tie-breaker for this specific pair

        # שיקום/שדרוג indicators (strong)
        shikum_terms = (
            "החלפה" in pn or "החלפת" in pn or
            "שיקום" in pn or "שדרוג" in pn or
            "חידוש" in pn or "חידושי" in pn or
            "שיפוץ" in pn or "תיקון" in pn or
            "איטום" in pn or "ציפוי" in pn or
            "שיפור" in pn or "חיזוק" in pn or
            "ישן" in pn or "קיימ" in pn  # "existing" indicators
        )

        # פיתוח indicators (strong)
        pituach_terms = (
            "הקמה" in pn or "הקמת" in pn or
            "חדש" in pn or "חדשה" in pn or "חדשים" in pn or
            "תוספת" in pn or "תוספות" in pn or
            "הרחבה" in pn or "הרחבת" in pn or
            "השלמת" in pn or "השלמה" in pn or  # Completing = new work
            "שלב ב" in pn or "שלב ג" in pn or "שלב 2" in pn or "שלב 3" in pn or
            "ותמל" in pn or 'ותמ"ל' in pn or "ותמ\"ל" in pn or
            "שכונה חדשה" in pn or "מתחם חדש" in pn or
            "פינוי בינוי" in pn or "תמא 38" in pn or
            "מתחם א" in pn or "מתחם ב" in pn  # Development phases
        )

        if shikum_terms and not pituach_terms:
            return "שיקום/שדרוג"
        if pituach_terms and not shikum_terms:
            return "פיתוח"
        # Both or neither - can't decide
        return None

    def _predict_by_fuzzy(pn: str, project_name: str) -> tuple[str | None, float, set[str] | None]:
        """
        Fuzzy/synonym-based prediction (Stage B).
        Returns: (label or None, confidence, candidate_allowed_set for LLM fallback)

        Logic:
        1. Try synonym matching to find subject
        2. Look up allowed_set for subject
        3. If single label -> return it
        4. If 2 labels (שיקום/שדרוג, פיתוח) -> apply tie-breaker
        5. If tie-breaker fails -> return allowed_set for LLM
        """
        pn_normalized = normalize_text(project_name)

        # ========================================
        # Step 1: Try synonym substring matching
        # ========================================
        subj = _subject_from_synonyms_normalized(pn_normalized)
        if subj:
            # Direct label match (e.g., קידוחים)
            if subj in ALLOWED_FUNDING_LABELS:
                return canonicalize_label(subj), 0.90, None

            # Look up allowed set for this subject
            allowed_set = SUBJECT_TO_ALLOWED.get(subj, set())

            if allowed_set and len(allowed_set) == 1:
                # Single option - deterministic
                return canonicalize_label(next(iter(allowed_set))), 0.90, None

            if allowed_set and len(allowed_set) == 2:
                # Try tie-breaker
                result = _apply_tiebreaker(pn, allowed_set)
                if result:
                    return result, 0.85, None
                # Tie-breaker failed - pass to LLM with narrowed set
                return None, 0.50, allowed_set

        # ========================================
        # Step 2: No synonym match - try fuzzy subject matching
        # ========================================
        subj, score = _best_subject_match(project_name)
        if not subj:
            # No subject found at all - try to classify by keywords alone
            # Default to קווים קצרים which allows {שיקום/שדרוג, פיתוח}
            default_allowed = {"שיקום/שדרוג", "פיתוח"}
            result = _apply_tiebreaker(pn, default_allowed)
            if result:
                return result, 0.60, None
            return None, 0.0, default_allowed

        allowed_set = SUBJECT_TO_ALLOWED.get(subj, set())
        if not allowed_set:
            return None, 0.0, None

        # ========================================
        # Step 3: Apply fuzzy threshold
        # ========================================
        if score >= R14_FUZZY_THRESHOLD:
            if len(allowed_set) == 1:
                return canonicalize_label(next(iter(allowed_set))), score, None

            # Multiple options - try tie-breaker
            result = _apply_tiebreaker(pn, allowed_set)
            if result:
                return result, score, None
            # Tie-breaker failed - pass to LLM
            return None, score, allowed_set

        # Below threshold - still try tie-breaker with lower confidence
        if len(allowed_set) == 2:
            result = _apply_tiebreaker(pn, allowed_set)
            if result:
                return result, 0.55, None
            return None, score, allowed_set

        return None, score, allowed_set

    def _predict_by_llm(project_name: str, allowed_set: set[str] | None) -> tuple[str | None, float]:
        """
        LLM-based prediction (Stage C).
        Returns: (label or None, confidence)

        IMPORTANT: Only call LLM when allowed_set == {שיקום/שדרוג, פיתוח}
        For other cases, deterministic classification should have succeeded.
        """
        # Only call LLM for the specific 2-way ambiguity case
        target_set = {"שיקום/שדרוג", "פיתוח"}
        if allowed_set != target_set:
            print(f"[R_14 DEBUG] LLM SKIPPED: allowed_set={allowed_set} != target {target_set}")
            return None, 0.0

        llm_enabled = bool(getattr(cfg, "llm_enabled", False))
        provider = getattr(cfg, "llm_provider", "gemini")
        provider_l = (provider or "").strip().lower()

        if provider_l == "gemini":
            has_key = bool(os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY"))
            default_model = "gemini-1.5-flash"
        else:
            has_key = bool(os.getenv("OPENAI_API_KEY"))
            default_model = "gpt-4o-mini"

        # TEMP DEBUG: Show why LLM is skipped or called
        print(f"[R_14 DEBUG] _predict_by_llm called: project={project_name[:30]!r}, llm_enabled={llm_enabled}, has_key={has_key}, provider={provider_l}")

        if not (llm_enabled and has_key):
            print(f"[R_14 DEBUG] LLM SKIPPED: llm_enabled={llm_enabled}, has_key={has_key}")
            return None, 0.0

        try:
            from .llm_client import classify_funding_with_confidence, LLMQuotaError

            model = getattr(cfg, "llm_model", default_model)

            # LLM only chooses between שיקום/שדרוג and פיתוח
            prompt = build_llm_prompt(project_name, allowed_set=target_set)
            print(f"[R_14 DEBUG] Calling LLM with model={model}, allowed_set={target_set}")
            llm_label, llm_conf = classify_funding_with_confidence(
                prompt, provider=provider_l, model=model
            )
            print(f"[R_14 DEBUG] LLM returned: label={llm_label!r}, conf={llm_conf}")

            return canonicalize_label(llm_label), float(llm_conf)

        except LLMQuotaError:
            # Quota exceeded - graceful fallback
            return None, 0.0
        except Exception as e:
            print(f"[R_14] LLM error (provider={provider_l}): {e!r}")
            return None, 0.0

    # ---------------------------
    # Column resolution
    # ---------------------------
    col_map = {_norm_col(c): c for c in report_df.columns}
    project_col = col_map.get("שם פרויקט")
    class_col = col_map.get("סיווג פרויקט")

    # Column H: "חומר מבנה" (material) — for forbidden name check
    _r14_material_col = col_map.get("חומר מבנה")
    _r14_material_col_idx = None
    if _r14_material_col is not None:
        try:
            _r14_material_col_idx = list(report_df.columns).index(_r14_material_col)
        except ValueError:
            _r14_material_col_idx = None

    missing = []
    if project_col is None:
        missing.append("שם פרויקט")
    if class_col is None:
        missing.append("סיווג פרויקט")

    if missing:
        return [
            CheckResult(
                rule_id=rule_id,
                rule_name=rule_name,
                severity=Severity.CRITICAL,
                sheet_name=sheet,
                row_index=None,
                column_name=None,
                key_context="columns_presence",
                actual_value=[_norm_col(c) for c in report_df.columns],
                expected_value=f"must include {sorted(['שם פרויקט', 'סיווג פרויקט'])}",
                status=Status.FAIL,
                message=f"חסרות עמודות (אחרי נירמול כותרות): {missing}",
            )
        ]

    # Excel cell mapping for row/column tracking
    from openpyxl.utils import get_column_letter as _get_col_letter
    _col_to_letter = {col: _get_col_letter(i + 1) for i, col in enumerate(report_df.columns)}
    _header_row = getattr(cfg, "report_header_row", 6)
    _excel_row_offset = _header_row + 3  # same convention as R12

    def _r14_cell_ref(df_i: int, col_name: str) -> Optional[str]:
        letter = _col_to_letter.get(col_name)
        if not letter:
            return None
        return f"{sheet}!{letter}{int(df_i + _excel_row_offset)}"

    # Row filtering setup
    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")
    id_col = col_map.get(id_norm)

    def _is_real_row(idx: int) -> bool:
        if not id_col:
            return True
        v = report_df.at[idx, id_col]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return False
        s = str(v).strip()
        if s in {"-", "", "nan", "None", "none"}:
            return False
        return True

    # Config flag for שיקום/שדרוג special treatment
    shikum_not_investment = bool(getattr(cfg, "r14_shikum_not_investment", False))

    # ---------------------------
    # Stage B3: Prior fallback from reported distribution
    # ---------------------------
    # Templates for bucketing generic project names (ordered by priority)
    PRIOR_TEMPLATES = [
        # Core infra patterns
        ("קווי מים", "קווי מים"),
        ("קו מים", "קו מים"),
        ("קווי ביוב", "קווי ביוב"),
        ("קו ביוב", "קו ביוב"),
        ("מערכת מים", "מערכת מים"),
        ("מערכת ביוב", "מערכת ביוב"),
        ("מאסף ביוב", "מאסף ביוב"),
        ("מאסף", "מאסף"),
        ("קו אספקה", "קו אספקה"),
        ("אספקת מים", "אספקת מים"),
        # Facilities
        ("תחנת שאיבה", "תחנת שאיבה"),
        ("תחנת בוסטרים", "בוסטרים"),
        ("בוסטר", "בוסטרים"),
        ("תחנת שאיבת מים", "תחנת שאיבת מים"),
        ("תחנת שאיבה", "תחנת שאיבה"),
        ("מאגר", "מאגר"),
        ("מערכת בקרה", "בקרה"),
        ("פיקוד ובקרה", "בקרה"),
        ("מיגון", "בקרה"),
        ("מגוף", "מגוף"),
        # Development zones (tend toward פיתוח)
        ('אז"ת', "אזור תעשייה"),
        ("אזור תעשייה", "אזור תעשייה"),
        ("א.ת", "אזור תעשייה"),
        ("היי-טק", "היי-טק"),
        ("הייטק", "היי-טק"),
        ("מתחם", "מתחם"),
        # Generic location patterns (most common, order matters - specific first)
        ("בשכונת", "שכונה"),  # More specific first
        ("בשכונה", "שכונה"),
        ("שכונת", "שכונה"),
        ("שכונה", "שכונה"),
        ("באזורים שונים", "אזורים"),
        ("במוקדים שונים", "אזורים"),
        ("אזורים מפוזרים", "אזורים"),
        ("אזורים שונים", "אזורים"),
        ("באזור", "אזורים"),
        # Road/highway patterns
        ("בכביש", "כביש"),
        ("כביש", "כביש"),
        ("ציר", "ציר"),
        ("ברחוב", "רחוב"),
        # Municipal patterns
        ("גרעין הישוב", "גרעין"),
        ("מרכז הישוב", "מרכז"),
        ("במרכז הישוב", "מרכז"),
        ("הסעת המונים", "הסעת המונים"),
        # Direction patterns
        ("צפוני", "כיוון"),
        ("דרומי", "כיוון"),
        ("מזרחי", "כיוון"),
        ("מערבי", "כיוון"),
        ("לכיוון", "כיוון"),
    ]

    PRIOR_MIN_SUPPORT = 5   # Minimum rows to learn prior (lowered from 8)
    PRIOR_MIN_SKEW = 0.67   # Minimum ratio for strong prior (lowered from 0.75)

    def _match_template(pn_norm: str) -> str | None:
        """Match project name to first matching template."""
        for pattern, template_name in PRIOR_TEMPLATES:
            if pattern in pn_norm:
                return template_name
        return None

    def _has_infra_or_action_token(pn_norm: str) -> bool:
        """
        Check if name has infrastructure or action tokens (not just location).
        This prevents classifying names like "שכונה צפונית" as location-only
        when they clearly have infra/action terms.
        """
        # Infrastructure tokens
        infra_tokens = [
            "קו", "קווי", "מערכת", "מים", "ביוב", "מאסף", "מאספים",
            "תחנה", "מאגר", "בריכ", "משאב", "סניקה", "בוסטר", "בוסטרים",
            'אז"ת', "אזור תעשי", "היי-טק", "הייטק", "מתחם",
            "מגוף", "מגופים", "מגוף מכני",
            'תש"ב', "תשב",  # תחנת שאיבה abbreviation
            "ותמ\"ל", 'ותמ"ל', "ותמל",  # Development projects
            "צנרת", "רשת",
        ]
        # Action tokens (שיקום/שדרוג or פיתוח indicators)
        action_tokens = [
            "שדרוג", "שיקום", "החלפ", "חידוש", "שיפוץ", "תיקון",
            "איטום", "ציפוי",  # Rehabilitation
            "הקמ", "חדש", "תוספ", "הרחב", "בנייה", "בניה",  # Development
        ]
        all_tokens = infra_tokens + action_tokens
        return any(t in pn_norm for t in all_tokens)

    # Compute priors from reported values (once per run)
    template_counts: dict[str, dict[str, int]] = {}  # template -> {label: count}
    target_labels = {"שיקום/שדרוג", "פיתוח"}

    for idx, row in report_df.iterrows():
        if not _is_real_row(idx):
            continue
        pn_raw = str(row.get(project_col, "")).strip()
        reported_raw = str(row.get(class_col, "")).strip()
        if not pn_raw or pn_raw.lower() == "nan":
            continue
        reported_label = canonicalize_label(reported_raw)
        if reported_label not in target_labels:
            continue
        pn_norm = normalize_text(pn_raw)
        template = _match_template(pn_norm)
        if template:
            if template not in template_counts:
                template_counts[template] = {"שיקום/שדרוג": 0, "פיתוח": 0}
            template_counts[template][reported_label] += 1

    def _get_prior(pn_norm: str) -> tuple[str | None, float]:
        """
        Get prior prediction based on template distribution.
        Returns (label, conf) or (None, 0).

        Uses tiered confidence:
        - Strong skew (>=67%) + good support (>=5): conf=0.60
        - Moderate skew (>=55%) + some support (>=3): conf=0.50
        - Weak majority (>50%) + minimal support (>=2): conf=0.45
        """
        template = _match_template(pn_norm)
        if not template or template not in template_counts:
            return None, 0.0
        counts = template_counts[template]
        total = counts["שיקום/שדרוג"] + counts["פיתוח"]
        if total < 2:  # Need at least 2 examples
            return None, 0.0

        winner = "שיקום/שדרוג" if counts["שיקום/שדרוג"] >= counts["פיתוח"] else "פיתוח"
        ratio = counts[winner] / total

        # Tiered confidence based on support and skew
        if total >= PRIOR_MIN_SUPPORT and ratio >= PRIOR_MIN_SKEW:
            return winner, 0.60  # Strong prior
        if total >= 3 and ratio >= 0.55:
            return winner, 0.50  # Moderate prior
        if total >= 2 and ratio > 0.50:
            return winner, 0.45  # Weak prior
        # Even split - can't decide
        return None, 0.0

    # ---------------------------
    # Main loop
    # ---------------------------
    results: List[CheckResult] = []

    for idx, row in report_df.iterrows():
        if not _is_real_row(idx):
            continue

        project_name = str(row.get(project_col, "")).strip()
        if not project_name or project_name.lower() == "nan":
            continue

        reported_raw = str(row.get(class_col, "")).strip()

        # --- Check 1: Missing reported value ---
        if _is_empty(reported_raw):
            msg = (
                "נכשל - ערך חסר בעמודת סיווג פרויקט\n"
                "סיווג סופי על ידי keyword: (אין)"
            )
            results.append(
                CheckResult(
                    rule_id=rule_id,
                    rule_name=rule_name,
                    severity=Severity.WARNING,
                    sheet_name=sheet,
                    row_index=idx,
                    column_name="סיווג פרויקט",
                    status=Status.FAIL,
                    message=msg,
                    actual_value="(ריק)",
                    expected_value=" | ".join(sorted(ALLOWED_FUNDING_LABELS)),
                    key_context=f"project_name={project_name}",
                )
            )
            continue

        reported = canonicalize_label(reported_raw)

        # --- Check 2: Illegal reported value ---
        if reported and (reported not in ALLOWED_FUNDING_LABELS):
            msg = (
                f"נכשל - ערך לא חוקי: '{reported_raw}'\n"
                "סיווג סופי על ידי keyword: (אין)"
            )
            results.append(
                CheckResult(
                    rule_id=rule_id,
                    rule_name=rule_name,
                    severity=Severity.WARNING,
                    sheet_name=sheet,
                    row_index=idx,
                    column_name="סיווג פרויקט",
                    status=Status.FAIL,
                    message=msg,
                    actual_value=reported_raw,
                    expected_value=" | ".join(sorted(ALLOWED_FUNDING_LABELS)),
                    key_context=f"project_name={project_name}",
                )
            )
            continue

        pn = normalize_text(project_name)
        word_count = len(pn.split())

        # --- Step 1: Forbidden Names ---
        if getattr(cfg, "street_lookup_enabled", True) and utility_name:
            forbidden_reason = None
            if "אחזקה" in pn:
                forbidden_reason = "שם מכיל 'אחזקה'"
            elif "שותף" in pn:
                forbidden_reason = "שם מכיל 'שותף'"
            if forbidden_reason is None and _r14_material_col is not None:
                try:
                    mat_raw = report_df.iloc[idx, _r14_material_col_idx] if _r14_material_col_idx is not None else None
                    if mat_raw is not None and not _is_empty(mat_raw):
                        mat_norm = normalize_text(str(mat_raw).strip())
                        if len(mat_norm) >= 2 and mat_norm in pn:
                            forbidden_reason = f"שם מכיל חומר מבנה '{mat_raw}'"
                except Exception:
                    pass
            if forbidden_reason:
                trace = f"method=forbidden | reason={forbidden_reason} | decision=שם פרויקט אסור"
                print(f"[R_14 DEBUG]   -> Forbidden: {forbidden_reason}")
                results.append(
                    CheckResult(
                        rule_id=rule_id,
                        rule_name=rule_name,
                        severity=Severity.WARNING,
                        sheet_name=sheet,
                        row_index=idx,
                        column_name="סיווג פרויקט",
                        status=Status.FORBIDDEN_PROJECT_NAME,
                        message=f"שם פרויקט אסור: {forbidden_reason}\n{trace}",
                        actual_value=reported_raw,
                        expected_value="(שם פרויקט אסור)",
                        key_context=f"project_name={project_name}",
                        confidence=1.0,
                        method="forbidden",
                    )
                )
                continue

        # --- Pipeline: Predict label ---
        predicted = None
        method = None
        confidence = 0.0
        candidate_allowed_set = None
        llm_was_attempted = False  # Track if LLM was actually called
        street_found = False  # Set True when street validation finds a match (any allow_clean outcome)

        # Step 2: Keyword
        kw_label, kw_conf = _predict_by_keyword(pn)
        # TEMP DEBUG
        print(f"[R_14 DEBUG] Row {idx}: project={project_name[:40]!r}")
        print(f"[R_14 DEBUG]   Keyword stage: label={kw_label!r}, conf={kw_conf}")
        if kw_label is not None:
            predicted = canonicalize_label(kw_label)
            method = "keyword"
            confidence = kw_conf

        # Step 3: Fuzzy (only if keyword didn't decide)
        if predicted is None:
            fuzzy_label, fuzzy_conf, fuzzy_allowed = _predict_by_fuzzy(pn, project_name)
            # TEMP DEBUG
            print(f"[R_14 DEBUG]   Fuzzy stage: label={fuzzy_label!r}, conf={fuzzy_conf}, allowed={fuzzy_allowed}")
            if fuzzy_label is not None:
                predicted = canonicalize_label(fuzzy_label)
                method = "fuzzy"
                confidence = fuzzy_conf
            else:
                candidate_allowed_set = fuzzy_allowed

        # Step 4: Street Validation (moved up). If street found with high confidence -> CLEAN/PASS.
        if predicted is None and getattr(cfg, "street_lookup_enabled", True) and utility_name:
            try:
                from .ckan_client import (
                    fetch_streets_for_city, find_best_street_match,
                    utility_to_city_name, is_short_name,
                    has_explicit_street_pattern, is_street_then_number,
                    PASSING_SCORE,
                )
                city_name = utility_to_city_name(utility_name)
                city_streets = fetch_streets_for_city(
                    city_name,
                    timeout=getattr(cfg, "street_lookup_timeout_sec", 10.0),
                )
                match = find_best_street_match(pn, city_streets)
                trace_parts = [
                    f"method=street_api",
                    f"city_norm={city_name}",
                    f"best_candidate={match.candidate or 'none'}",
                    f"best_score={match.score}",
                    f"match_type={match.match_type or 'none'}",
                ]
                if match.found:
                    street_found = True
                    # Exact match or high-confidence score → CLEAN (skip LLM); else require short/explicit pattern
                    allow_clean = (
                        is_short_name(project_name)
                        or has_explicit_street_pattern(project_name)
                        or is_street_then_number(project_name)
                        or getattr(match, "exact_match", False)
                        or (match.score >= PASSING_SCORE)
                    )
                    if allow_clean:
                        trace_parts.append("decision=לא חשוד")
                        trace = " | ".join(trace_parts)
                        print(f"[R_14 DEBUG]   -> Street lookup: CLEAN ('{match.street}' in {city_name}, score={match.score})")
                        results.append(
                            CheckResult(
                                rule_id=rule_id,
                                rule_name=rule_name,
                                severity=Severity.INFO,
                                sheet_name=sheet,
                                row_index=idx,
                                column_name="סיווג פרויקט",
                                status=Status.CLEAN,
                                message=f"שם רחוב תקין - נמצא '{match.street}' ברשימת הרחובות של {city_name}\n{trace}",
                                actual_value=reported_raw,
                                expected_value=f"רחוב מאומת: {match.street}",
                                key_context=f"project_name={project_name}",
                                confidence=1.0,
                                method="street_lookup",
                            )
                        )
                        continue
                    else:
                        trace_parts.append("decision=skip_long_name")
                        print(f"[R_14 DEBUG]   -> Street found '{match.street}' but name too long, keeping INFO")
                else:
                    trace_parts.append("decision=שם פרויקט אסור")
                    trace = " | ".join(trace_parts)
                    print(f"[R_14 DEBUG]   -> Street lookup: FORBIDDEN (no match in {city_name}, score={match.score})")
                    results.append(
                        CheckResult(
                            rule_id=rule_id,
                            rule_name=rule_name,
                            severity=Severity.WARNING,
                            sheet_name=sheet,
                            row_index=idx,
                            column_name="סיווג פרויקט",
                            status=Status.FORBIDDEN_PROJECT_NAME,
                            message=f"שם פרויקט אסור - לא נמצא רחוב תואם ברשימת הרחובות של {city_name}\n{trace}",
                            actual_value=reported_raw,
                            expected_value="(שם רחוב לא מזוהה)",
                            key_context=f"project_name={project_name}",
                            confidence=1.0,
                            method="street_lookup",
                        )
                    )
                    continue
            except Exception as e:
                print(f"[R_14 DEBUG]   -> Street lookup error: {e!r}")

        # Step 5: Structural Guard — do not call LLM for generic short names without street match.
        if predicted is None and _is_generic_project_name(project_name) and not street_found:
            print(f"[R_14 DEBUG]   -> Structural guard: generic name rejected (no street match), skipping LLM")
            msg = (
                "לא ניתן לסווג - שם כללי נחסם על ידי structural guard\n"
                "Generic name rejected by structural guard"
            )
            results.append(
                CheckResult(
                    rule_id=rule_id,
                    rule_name=rule_name,
                    severity=Severity.INFO,
                    sheet_name=sheet,
                    row_index=idx,
                    column_name="סיווג פרויקט",
                    status=Status.INFO,
                    message=msg,
                    actual_value=reported_raw,
                    expected_value="(לא ניתן לחזות)",
                    key_context=f"project_name={project_name}",
                    confidence=0.0,
                    method="structural_guard",
                )
            )
            continue

        # Step 6: LLM (only if keyword + fuzzy failed and not blocked by structural guard)
        if predicted is None:
            print(f"[R_14 DEBUG]   -> Entering LLM stage (keyword+fuzzy failed)")
            target_set = {"שיקום/שדרוג", "פיתוח"}
            if candidate_allowed_set == target_set:
                llm_was_attempted = True
            llm_label, llm_conf = _predict_by_llm(project_name, candidate_allowed_set)
            if llm_label is not None:
                predicted = canonicalize_label(llm_label)
                method = "llm"
                confidence = llm_conf
        else:
            print(f"[R_14 DEBUG]   -> LLM SKIPPED (predicted already set by {method})")

        # Stage D: Prior fallback
        # Run when:
        # 1. Still no prediction AND
        # 2. Either allowed_set is the 2-way ambiguity OR name matches a strong template
        if predicted is None:
            use_prior = False
            # Case 1: Explicit 2-way ambiguity from fuzzy stage
            if candidate_allowed_set == {"שיקום/שדרוג", "פיתוח"}:
                use_prior = True
            # Case 2: Subject matching failed but name matches a strong infra template
            # (templates like מערכת/קווי/מאסף + מים/ביוב imply 2-way ambiguity)
            elif candidate_allowed_set is None or len(candidate_allowed_set or set()) == 0:
                template = _match_template(pn)
                if template is not None:
                    use_prior = True
                    print(f"[R_14 DEBUG]   -> Template matched ({template}), enabling prior fallback")

            if use_prior:
                prior_label, prior_conf = _get_prior(pn)
                if prior_label is not None:
                    print(f"[R_14 DEBUG]   -> Prior fallback: label={prior_label!r}, conf={prior_conf}")
                    predicted = prior_label
                    method = "prior"
                    confidence = prior_conf

        # Stage E: Global fallback - use overall file distribution for 2-way ambiguity
        # Only when we have clear infra pattern but no template-specific prior
        if predicted is None and candidate_allowed_set == {"שיקום/שדרוג", "פיתוח"}:
            # Check if name has clear infra pattern
            has_clear_infra = any(t in pn for t in ["מערכת", "קווי", "קו ", "מאסף", "תחנ", "מאגר"])
            if has_clear_infra:
                # Use global distribution from template_counts
                global_shikum = sum(c.get("שיקום/שדרוג", 0) for c in template_counts.values())
                global_pituach = sum(c.get("פיתוח", 0) for c in template_counts.values())
                total = global_shikum + global_pituach
                if total >= 10:  # Need reasonable sample
                    dominant = "שיקום/שדרוג" if global_shikum >= global_pituach else "פיתוח"
                    ratio = max(global_shikum, global_pituach) / total
                    if ratio >= 0.55:  # At least 55% skew
                        print(f"[R_14 DEBUG]   -> Global fallback: {dominant} ({ratio:.0%} of {total})")
                        predicted = dominant
                        method = "prior"
                        confidence = 0.40  # Very low confidence

        # --- No decision (keyword/fuzzy/street/structural guard/LLM/prior all failed or unavailable) ---
        # Forbidden and street validation already run above; here we only emit no-decision.
        if predicted is None:
            if not _has_infra_or_action_token(pn):
                print(f"[R_14 DEBUG]   -> No prediction: location-only name")
                msg = (
                    "לא ניתן לסווג - שם פרויקט כללי/מיקום בלבד (חסר סוג תשתית/פעולה)\n"
                    "סיווג סופי: לא זמין"
                )
            else:
                print(f"[R_14 DEBUG]   -> No prediction available, showing as INFO")
                msg = (
                    "לא ניתן לסווג - keyword/fuzzy/LLM לא החזירו תוצאה\n"
                    "סיווג סופי: לא זמין"
                )
            no_decision_method = "fail_llm" if llm_was_attempted else "no_decision"
            results.append(
                CheckResult(
                    rule_id=rule_id,
                    rule_name=rule_name,
                    severity=Severity.INFO,
                    sheet_name=sheet,
                    row_index=idx,
                    column_name="סיווג פרויקט",
                    status=Status.INFO,
                    message=msg,
                    actual_value=reported_raw,
                    expected_value="(לא ניתן לחזות)",
                    key_context=f"project_name={project_name}",
                    confidence=0.0,
                    method=no_decision_method,
                )
            )
            continue

        # --- Short name street override (≤ 3 meaningful tokens) ---
        # If pipeline produced a prediction but name is short and matches a street,
        # override with CLEAN (e.g. "משה דיין" classified as "פיתוח" but it's just a street)
        if getattr(cfg, "street_lookup_enabled", True) and utility_name:
            try:
                from .ckan_client import (
                    fetch_streets_for_city, find_best_street_match,
                    utility_to_city_name, is_short_name,
                    has_explicit_street_pattern, is_street_then_number,
                    PASSING_SCORE,
                )
                _allow_override = (
                    is_short_name(project_name)
                    or has_explicit_street_pattern(project_name)
                    or is_street_then_number(project_name)
                )
                if _allow_override:
                    _city = utility_to_city_name(utility_name)
                    _city_streets = fetch_streets_for_city(
                        _city,
                        timeout=getattr(cfg, "street_lookup_timeout_sec", 10.0),
                    )
                    _match = find_best_street_match(pn, _city_streets)
                    if _match.found and (
                        getattr(_match, "exact_match", False) or _match.score >= PASSING_SCORE
                    ):
                        _trace = f"method=street_api | city_norm={_city} | best_candidate={_match.candidate} | best_score={_match.score} | match_type={_match.match_type} | decision=לא חשוד (override {method})"
                        print(f"[R_14 DEBUG]   -> Short name street override: '{project_name}' matched '{_match.street}' in {_city} (was {method}={predicted}, score={_match.score})")
                        results.append(
                            CheckResult(
                                rule_id=rule_id,
                                rule_name=rule_name,
                                severity=Severity.INFO,
                                sheet_name=sheet,
                                row_index=idx,
                                column_name="סיווג פרויקט",
                                status=Status.CLEAN,
                                message=f"שם רחוב תקין - נמצא '{_match.street}' ברשימת הרחובות של {_city}\n{_trace}",
                                actual_value=reported_raw,
                                expected_value=f"רחוב מאומת: {_match.street}",
                                key_context=f"project_name={project_name}",
                                confidence=1.0,
                                method="street_lookup",
                            )
                        )
                        continue
            except Exception as e:
                print(f"[R_14 DEBUG]   -> Short name street lookup error: {e!r}")

        # Ensure canonical forms
        predicted = canonicalize_label(predicted)
        reported = canonicalize_label(reported)

        # --- Hard rule: תחזוקה / שוטף is NOT investment ---
        if reported == "תחזוקה / שוטף":
            msg = (
                "נכשל - לפי ערך בפועל: אינו שייך לתוכנית השקעה\n"
                + _final_line(method, predicted, confidence)
            )
            results.append(
                CheckResult(
                    rule_id=rule_id,
                    rule_name=rule_name,
                    severity=Severity.WARNING,
                    sheet_name=sheet,
                    row_index=idx,
                    column_name="סיווג פרויקט",
                    status=Status.FAIL,
                    message=msg,
                    actual_value=reported_raw,
                    expected_value="(לא תחזוקה / שוטף)",
                    key_context=f"project_name={project_name}",
                    confidence=confidence,
                    method=method,
                )
            )
            continue

        if predicted == "תחזוקה / שוטף":
            msg = (
                "נכשל - לפי ערך צפוי: אינו שייך לתוכנית השקעה\n"
                + _final_line(method, predicted, confidence)
            )
            results.append(
                CheckResult(
                    rule_id=rule_id,
                    rule_name=rule_name,
                    severity=Severity.WARNING,
                    sheet_name=sheet,
                    row_index=idx,
                    column_name="סיווג פרויקט",
                    status=Status.FAIL,
                    message=msg,
                    actual_value=reported_raw,
                    expected_value="(לא תחזוקה / שוטף)",
                    key_context=f"project_name={project_name}",
                    confidence=confidence,
                    method=method,
                )
            )
            continue

        # --- Optional: שיקום/שדרוג NOT investment mode ---
        if shikum_not_investment and reported == "שיקום/שדרוג" and predicted == "שיקום/שדרוג":
            msg = (
                "נכשל - לפי דרישת לקוח: שיקום/שדרוג אינו שייך לתוכנית השקעה\n"
                + _final_line(method, predicted, confidence)
            )
            results.append(
                CheckResult(
                    rule_id=rule_id,
                    rule_name=rule_name,
                    severity=Severity.WARNING,
                    sheet_name=sheet,
                    row_index=idx,
                    column_name="סיווג פרויקט",
                    status=Status.FAIL,
                    message=msg,
                    actual_value=reported_raw,
                    expected_value="(לא שיקום/שדרוג)",
                    key_context=f"project_name={project_name}",
                    confidence=confidence,
                    method=method,
                )
            )
            continue

        # --- Compare reported vs predicted ---
        LOW_CONF_THRESHOLD = 0.70  # Below this = low confidence (gray/INFO)

        # Special note for prior-based predictions
        prior_note = ""
        if method == "prior":
            prior_note = "\n(סיווג בוצע לפי התפלגות ערכים מדווחים בתאגיד עבור תבנית שם דומה. ייתכן שהשם כללי מדי.)"

        if reported == predicted:
            # Check if low confidence match
            if confidence < LOW_CONF_THRESHOLD and method != "keyword":
                # Low confidence PASS - show as PASS with INFO severity (gray)
                msg = f"עבר (ביטחון נמוך): התאמה\n{_final_line(method, predicted, confidence)}{prior_note}"
                results.append(
                    CheckResult(
                        rule_id=rule_id,
                        rule_name=rule_name,
                        severity=Severity.INFO,
                        sheet_name=sheet,
                        row_index=idx,
                        column_name="סיווג פרויקט",
                        status=Status.PASS_,  # Show עבר but with gray styling
                        message=msg,
                        actual_value=reported_raw,
                        expected_value=predicted,
                        key_context=f"project_name={project_name}",
                        confidence=confidence,
                        method=method,
                    )
                )
            else:
                # High confidence PASS (prior always has low conf, so won't reach here)
                msg = f"עבר: התאמה\n{_final_line(method, predicted, confidence)}{prior_note}"
                results.append(
                    CheckResult(
                        rule_id=rule_id,
                        rule_name=rule_name,
                        severity=Severity.INFO,
                        sheet_name=sheet,
                        row_index=idx,
                        column_name="סיווג פרויקט",
                        status=Status.PASS_,
                        message=msg,
                        actual_value=reported_raw,
                        expected_value=predicted,
                        key_context=f"project_name={project_name}",
                        confidence=confidence,
                        method=method,
                    )
                )
        else:
            # FAIL - mismatch
            if confidence < LOW_CONF_THRESHOLD and method != "keyword":
                # Low confidence FAIL - show as FAIL with INFO severity (gray)
                msg = (
                    "אי-התאמה (ביטחון נמוך)\n"
                    + _final_line(method, predicted, confidence)
                    + prior_note
                )
                results.append(
                    CheckResult(
                        rule_id=rule_id,
                        rule_name=rule_name,
                        severity=Severity.INFO,
                        sheet_name=sheet,
                        row_index=idx,
                        column_name="סיווג פרויקט",
                        status=Status.FAIL,  # Show נכשל but with gray styling
                        message=msg,
                        actual_value=reported_raw,
                        expected_value=predicted,
                        key_context=f"project_name={project_name}",
                        confidence=confidence,
                        method=method,
                    )
                )
            else:
                # High confidence FAIL (prior always has low conf, so won't reach here)
                msg = (
                    "נכשל - סיווג לא תואם\n"
                    + _final_line(method, predicted, confidence)
                    + prior_note
                )
                results.append(
                    CheckResult(
                        rule_id=rule_id,
                        rule_name=rule_name,
                        severity=Severity.WARNING,
                        sheet_name=sheet,
                        row_index=idx,
                        column_name="סיווג פרויקט",
                        status=Status.FAIL,
                        message=msg,
                        actual_value=reported_raw,
                        expected_value=predicted,
                        key_context=f"project_name={project_name}",
                        confidence=confidence,
                        method=method,
                    )
                )

    # Post-process: add excel_cells to all R14 results for row/column tracking
    for r in results:
        if r.row_index is not None and r.excel_cells is None:
            ref = _r14_cell_ref(r.row_index, class_col)
            if ref:
                r.excel_cells = [ref]

    return results



INVALID_PROJECT_TEXTS = {
    "רחוב",
    "בין הבתים",
    "שטח פתוח",
    "רחוב שכונת",
    "רחוב שכונה",   # practical variant
}

def check_015_invalid_project_names(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    R_15 — שמות פרויקטים לא תקינים (raw row level).

    One CheckResult per row. For each real row, checks "שם פרויקט" and "מיקום פרויקט":
    values must not be generic ("רחוב", "בין הבתים", "שטח פתוח", "רחוב שכונת" etc.).
    Skips location-helper text "רחוב/שכונה/תב\"ע". Emits PASS or FAIL per row.
    """



    def _norm_col(col_name: object) -> str:
        """
        Normalize Excel column headers for robust matching.
        Handles trailing spaces, newlines, NBSP, repeated whitespace,
        and ignores helper header text like 'רחוב/שכונה/תב"ע'.
        Also removes pandas duplicate suffix like ".1".
        """
        s = "" if col_name is None else str(col_name)

        # normalize whitespace-ish characters
        s = s.replace("\u00A0", " ")   # NBSP -> space
        s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")

        # ignore helper header fragment that appears due to merged cells
        s = s.replace('רחוב/שכונה/תב"ע', "")
        s = s.replace("רחוב/שכונה/תב״ע", "")

        # collapse whitespace
        s = " ".join(s.split()).strip()

        # remove pandas duplicate suffix: ".1", ".2", ...
        s = re.sub(r"\.\d+$", "", s).strip()
        return s


    RULE_ID = "R_15"
    RULE_NAME = "שמות פרויקטים לא תקינים"
    SHEET = getattr(cfg, "report_sheet_name", "גיליון דיווח")

    # ---- Patterns ----
    INVALID_PROJECT_TEXTS = {
        normalize_text("רחוב"),
        normalize_text("בין הבתים"),
        normalize_text("שטח פתוח"),
        normalize_text("רחוב שכונת"),
        normalize_text("רחוב שכונה"),
        normalize_text('רחוב שכונת"'),
    }

    # אם תרצה הרחבה בהמשך (אופציונלי)
    INVALID_PROJECT_REGEXES = [
        re.compile(r"^רחוב\s*$"),
        re.compile(r"^בין\s+הבתים\s*$"),
        re.compile(r"^שטח\s+פתוח\s*$"),
        re.compile(r"^רחוב\s+שכונ(ה|ת)\s*$"),
    ]

    LOCATION_HELPER_TEXTS = {
        'רחוב/שכונה/תב"ע',
        'רחוב/שכונה/תב״ע',
        'רחוב/שכונה/תב”ע',
    }

    results: List[CheckResult] = []

    # ---- Resolve columns robustly (handles Unnamed + spaces + 2-row headers) ----
    norm_to_orig = {_norm_col(c): c for c in report_df.columns}

    def _tokens(s: str) -> set[str]:
        return {t for t in _norm_col(s).split(" ") if t}

    def _resolve_required(required: str) -> str | None:
        req_norm = _norm_col(required)
        req_tokens = _tokens(required)

        # 1) exact normalized match
        if req_norm in norm_to_orig:
            return norm_to_orig[req_norm]

        # 2) token containment (best for headers with suffixes like "שם פרויקט Unnamed...")
        candidates = []
        for norm_name, orig_name in norm_to_orig.items():
            cand_tokens = _tokens(norm_name)
            if req_tokens.issubset(cand_tokens):
                candidates.append((len(cand_tokens), len(norm_name), orig_name))

        if candidates:
            candidates.sort()
            return candidates[0][2]

        # 3) substring fallback
        for norm_name, orig_name in norm_to_orig.items():
            if req_norm and (req_norm in norm_name or norm_name in req_norm):
                return orig_name

        return None

    required_logical = ["שם פרויקט", "מיקום פרויקט"]
    resolved_map: dict[str, str] = {}
    missing_cols: list[str] = []

    for logical in required_logical:
        resolved = _resolve_required(logical)
        if resolved is None:
            missing_cols.append(logical)
        else:
            resolved_map[logical] = resolved

    if missing_cols:
        results.append(
            CheckResult(
                rule_id=RULE_ID,
                rule_name=RULE_NAME,
                severity=Severity.WARNING,
                sheet_name=SHEET,
                status=Status.FAIL,
                message=f"חסרות עמודות נדרשות לבדיקה: {', '.join(missing_cols)}",
                row_index=None,
                column_name=" / ".join(missing_cols),
                key_context="columns_presence",
            )
        )
        return results

    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")
    id_col = _resolve_required(id_norm)
    header_row = getattr(cfg, "report_header_row", 6)
    excel_row_offset = header_row + 3

    def _is_empty(v: object) -> bool:
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none"}

    def _is_real_row(df_i: int) -> bool:
        if id_col is None:
            return True
        v = report_df.at[df_i, id_col]
        if _is_empty(v):
            return False
        s = str(v).strip()
        return s not in {"-", ""} and s.lower() != "nan"

    def _is_invalid_value(val_str: str) -> bool:
        if not val_str or val_str.lower() in {"nan", "none"}:
            return False
        norm = normalize_text(val_str)
        return (norm in INVALID_PROJECT_TEXTS) or any(rx.match(norm) for rx in INVALID_PROJECT_REGEXES)

    for i in range(len(report_df)):
        if not _is_real_row(i):
            continue
        excel_row = int(i + excel_row_offset)
        pid = str(report_df.at[i, id_col]).strip() if id_col else ""
        errors: List[str] = []
        for logical_col, actual_df_col in resolved_map.items():
            raw_val = report_df.at[i, actual_df_col]
            if raw_val is None:
                continue
            val_str = str(raw_val).strip()
            if val_str == "" or val_str.lower() in {"nan", "none"}:
                continue
            if logical_col == "מיקום פרויקט" and val_str in LOCATION_HELPER_TEXTS:
                continue
            if _is_invalid_value(val_str):
                errors.append(f"{logical_col}: ערך כללי מדי ({val_str})")
        if not errors:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.PASS_,
                    message="תקין",
                    row_index=excel_row,
                    column_name=None,
                    key_context=f"ID={pid}",
                    actual_value="תקין",
                    expected_value="תיאור פרויקט מפורט",
                )
            )
        else:
            msg = "נכשל: " + "; ".join(errors)
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    message=msg,
                    row_index=excel_row,
                    column_name="שם פרויקט / מיקום פרויקט",
                    key_context=f"ID={pid}",
                    actual_value="; ".join(errors),
                    expected_value='תיאור מפורט (לא "רחוב"/"בין הבתים"/"שטח פתוח"/"רחוב שכונה")',
                )
            )

    return results

def check_018_facility_rehab_upgrade(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    R_18 — בדיקת שיקום ושדרוג מתקנים

    One CheckResult per valid row (after gates). Gatekeeper: F = "שיקום ושדרוג" AND
    G contains "מתקני מים", "מתקני ביוב", or "קידוחים". Aggregated validations:
    N (1960..current_year), O and P non-empty, at least one X in Q–U. Errors collected
    and joined with "; "; PASS message "פרויקט תקין", FAIL message "נכשל: ...".
    """
    from datetime import datetime
    from openpyxl.utils import get_column_letter, column_index_from_string

    RULE_BASE = "R_18"
    RULE_NAME = "בדיקת שיקום ושדרוג מתקנים"
    SHEET = getattr(cfg, "report_sheet_name", "גיליון דיווח")

    MIN_YEAR = 1960
    CUR_YEAR = datetime.now().year

    cols = list(report_df.columns)

    def _col_by_excel_letter(letter: str) -> Optional[str]:
        idx_1b = column_index_from_string(letter)
        idx_0b = idx_1b - 1
        if idx_0b < 0 or idx_0b >= len(cols):
            return None
        return cols[idx_0b]

    def _norm_col(c: object) -> str:
        s = str(c) if c is not None else ""
        s = re.sub(r"\s*Unnamed:.*", "", s)
        s = re.sub(r"_level_\d+", "", s)
        s = s.replace("\n", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    norm_to_orig: dict[str, str] = {}
    for c in cols:
        n = _norm_col(c)
        if n and n not in norm_to_orig:
            norm_to_orig[n] = c

    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")
    id_col = norm_to_orig.get(id_norm)

    col_f = _col_by_excel_letter("F")
    col_g = _col_by_excel_letter("G")
    col_n = _col_by_excel_letter("N")   # שנת הקמה
    col_o = _col_by_excel_letter("O")   # סוג מתקן
    col_p = _col_by_excel_letter("P")   # נפח/ספיקה
    cols_q_to_u = [_col_by_excel_letter(letter) for letter in "QRSTU"]

    missing = []
    if id_col is None:
        missing.append(id_norm)
    if col_f is None:
        missing.append("F")
    if col_g is None:
        missing.append("G")
    if col_n is None:
        missing.append("N")
    if col_o is None:
        missing.append("O")
    if col_p is None:
        missing.append("P")
    if not all(cols_q_to_u):
        missing.append("Q-U")

    if missing:
        return [
            CheckResult(
                rule_id=f"{RULE_BASE}_מבנה עמודות",
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                message=f"Missing required columns: {missing}",
                actual_value=list(report_df.columns),
                expected_value=["F", "G", "N", "O", "P", "Q-U"],
                row_index=None,
                column_name=None,
                key_context="columns_presence",
            )
        ]

    col_to_excel_letter = {col: get_column_letter(i + 1) for i, col in enumerate(cols)}
    header_row = getattr(cfg, "report_header_row", 6)
    excel_row_offset = header_row + 3

    def _cell_ref(df_i: int, col_name: str) -> Optional[str]:
        letter = col_to_excel_letter.get(col_name)
        if not letter:
            return None
        return f"{SHEET}!{letter}{int(df_i + excel_row_offset)}"

    def _is_empty(v: object) -> bool:
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        if isinstance(v, str) and v.strip() == "":
            return True
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none"}

    def _is_real_row(df_i: int) -> bool:
        if id_col is None:
            return True
        v = report_df.at[df_i, id_col]
        if _is_empty(v):
            return False
        s = str(v).strip()
        return s not in {"-", ""} and s.lower() != "nan"

    results: List[CheckResult] = []
    seen_project_ids: set[str] = set()

    for i in range(len(report_df)):
        if not _is_real_row(i):
            continue

        pid = str(report_df.at[i, id_col]).strip()
        if pid in seen_project_ids:
            continue
        seen_project_ids.add(pid)
        excel_row = int(i + excel_row_offset)

        raw_f = str(report_df.at[i, col_f] if col_f else "").strip()
        f_norm = normalize_text(raw_f)
        f_ok = f_norm in (normalize_text("שיקום ושדרוג"), normalize_text("שיקום/שדרוג"))
        if not f_ok:
            results.append(
                CheckResult(
                    rule_id=RULE_BASE,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    message=f"לא רלוונטי: סיווג פרויקט אינו שיקום ושדרוג (F={raw_f or '(ריק)'})",
                    row_index=excel_row,
                    column_name="F",
                    key_context=f"ID={pid}",
                    actual_value=raw_f or "(ריק)",
                    expected_value="שיקום ושדרוג",
                )
            )
            continue

        raw_g = str(report_df.at[i, col_g] if col_g else "").strip()
        g_norm = normalize_text(raw_g)
        g_ok = "מתקני מים" in g_norm or "מתקני ביוב" in g_norm or "קידוחים" in g_norm
        if not g_ok:
            results.append(
                CheckResult(
                    rule_id=RULE_BASE,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    message=f"לא רלוונטי: קוד הנדסי אינו מתקני מים / מתקני ביוב / קידוחים (G={raw_g or '(ריק)'})",
                    row_index=excel_row,
                    column_name="G",
                    key_context=f"ID={pid}",
                    actual_value=raw_g or "(ריק)",
                    expected_value="מתקני מים / מתקני ביוב / קידוחים",
                )
            )
            continue

        # --- Row-level validations: aggregate into errors[] then one CheckResult ---
        errors: List[str] = []

        raw_year = report_df.at[i, col_n]
        year_ok = True
        if _is_empty(raw_year):
            year_ok = False
        else:
            try:
                year_int = int(float(str(raw_year).strip()))
                if year_int < MIN_YEAR or year_int > CUR_YEAR:
                    year_ok = False
            except Exception:
                year_ok = False
        if not year_ok:
            errors.append("שנת הקמה לא תקינה או חסרה")

        if _is_empty(report_df.at[i, col_o]):
            errors.append("עמודה O צריכה להיות מלאה")

        if _is_empty(report_df.at[i, col_p]):
            errors.append("עמודה P צריכה להיות מלאה")

        has_x = False
        for dc in cols_q_to_u:
            if dc is None:
                continue
            v = report_df.at[i, dc]
            if not _is_empty(v) and str(v).strip().upper() == "X":
                has_x = True
                break
        if not has_x:
            errors.append("צריך להיות מסומן X אחד לפחות בעמודות Q עד U")

        if not errors:
            results.append(
                CheckResult(
                    rule_id=RULE_BASE,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.PASS_,
                    message="פרויקט תקין",
                    row_index=excel_row,
                    column_name=None,
                    key_context=f"ID={pid}",
                    actual_value="תקין",
                    expected_value="N,O,P תקינים; X ב-Q–U",
                )
            )
        else:
            msg = "נכשל: " + "; ".join(errors)
            refs = []
            for dc in [col_n, col_o, col_p] + [c for c in cols_q_to_u if c]:
                if dc:
                    r = _cell_ref(i, dc)
                    if r:
                        refs.append(r)
            results.append(
                CheckResult(
                    rule_id=RULE_BASE,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    message=msg,
                    row_index=excel_row,
                    column_name="N,O,P,Q-U",
                    key_context=f"ID={pid}",
                    actual_value="; ".join(errors),
                    expected_value="שנת הקמה תקינה; O,P מלאות; X ב-Q–U",
                    excel_cells=refs if refs else None,
                )
            )

    return results

def check_019_total_planned_cost_per_project(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    R_19 — סה"כ עלות מתוכננת לפרויקט

    Per unique 'מס' פרויקט' row:
      אומדן פרויקט >= (Excel columns X + AA + AD) in the same row.

    Notes:
      - X/AA/AD are taken by Excel letter positions (stable layout).
      - Emits PASS + FAIL per project id.
    """
    from openpyxl.utils import get_column_letter, column_index_from_string

    RULE_ID = "R_19"
    RULE_NAME = 'סה"כ עלות מתוכננת לפרויקט'
    SHEET = getattr(cfg, "report_sheet_name", "גיליון דיווח")

    cols = list(report_df.columns)

    def _is_empty(v: object) -> bool:
        if v is None:
            return True
        if pd.isna(v):
            return True
        if isinstance(v, str) and v.strip() == "":
            return True
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none"}

    def _to_number(v: object) -> Optional[float]:
        if _is_empty(v):
            return None
        try:
            s = str(v).strip()
            s = s.replace(",", "")
            return float(s)
        except Exception:
            return None

    def _fmt_int(v: Optional[float]) -> str:
        if v is None:
            return ""
        try:
            return str(int(round(float(v))))
        except Exception:
            return ""


    # Excel -> df column by letter (A=1)
    def _col_by_excel_letter(letter: str) -> Optional[str]:
        idx_1b = column_index_from_string(letter)  # A=1
        idx_0b = idx_1b - 1
        if idx_0b < 0 or idx_0b >= len(cols):
            return None
        return cols[idx_0b]

    # Resolve required columns
    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")
    estimate_key = "אומדן פרויקט"

    # robust header normalization (like R18/R15)
    def _norm_col(c: object) -> str:
        s = str(c) if c is not None else ""
        s = re.sub(r"\s*Unnamed:.*", "", s)      # remove Unnamed artifacts
        s = re.sub(r"_level_\d+", "", s)         # remove MultiIndex level suffix
        s = s.replace("\u00A0", " ")             # NBSP
        s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    norm_to_orig: dict[str, str] = {}
    for c in cols:
        n = _norm_col(c)
        if n and n not in norm_to_orig:
            norm_to_orig[n] = c

    def _tokens(s: str) -> set[str]:
        return {t for t in _norm_col(s).split(" ") if t}

    def _resolve_required(key: str) -> Optional[str]:
        key_norm = _norm_col(key)
        key_tokens = _tokens(key)

        # 1) exact match
        if key_norm in norm_to_orig:
            return norm_to_orig[key_norm]

        # 2) token containment (best for "אומדן פרויקט Unnamed: ...")
        candidates = []
        for norm_name, orig_name in norm_to_orig.items():
            cand_tokens = _tokens(norm_name)
            if key_tokens.issubset(cand_tokens):
                candidates.append((len(cand_tokens), len(norm_name), orig_name))

        if candidates:
            candidates.sort()
            return candidates[0][2]

        # 3) prefix / contains fallback
        for norm_name, orig_name in norm_to_orig.items():
            if norm_name.startswith(key_norm) or (key_norm in norm_name):
                return orig_name

        return None

    id_col = _resolve_required(id_norm)

    # Try by header name first, BUT fallback to Excel column AE (matches "AE >= X+AA+AD")
    estimate_col = _resolve_required(estimate_key) or _col_by_excel_letter("AE")



    x_col = _col_by_excel_letter("X")
    aa_col = _col_by_excel_letter("AA")
    ad_col = _col_by_excel_letter("AD")
    print("R19 resolved:", {"id": id_col, "estimate": estimate_col, "X": x_col, "AA": aa_col, "AD": ad_col})

    missing = []
    if id_col is None:
        missing.append(id_norm)
    if estimate_col is None:
        missing.append(f"{estimate_key} (or AE)")
    if x_col is None:
        missing.append("X")
    if aa_col is None:
        missing.append("AA")
    if ad_col is None:
        missing.append("AD")

    if missing:
        return [
            CheckResult(
                rule_id=f"{RULE_ID}_מבנה עמודות",
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name=None,
                key_context="columns_presence",
                actual_value=list(report_df.columns),
                expected_value=[id_norm, estimate_key, "X", "AA", "AD"],
                message=f"Missing required columns: {missing}",
            )
        ]

    # Excel highlighting mapping
    col_to_excel_letter = {col: get_column_letter(i + 1) for i, col in enumerate(cols)}
    header_row = getattr(cfg, "report_header_row", 6)
    excel_row_offset = header_row + 3  # same as R12/R18

    def _cell_ref(df_i: int, col_name: str) -> Optional[str]:
        letter = col_to_excel_letter.get(col_name)
        if not letter:
            return None
        excel_row = int(df_i + excel_row_offset)
        return f"{SHEET}!{letter}{excel_row}"

    def _is_real_row(df_i: int) -> bool:
        v = report_df.at[df_i, id_col]
        if _is_empty(v):
            return False
        s = str(v).strip()
        if s in {"-", ""}:
            return False
        if s.lower() == "nan":
            return False
        return True

    results: List[CheckResult] = []
    seen_project_ids: set[str] = set()

    for i in range(len(report_df)):
        if not _is_real_row(i):
            continue

        pid = str(report_df.at[i, id_col]).strip()
        if pid in seen_project_ids:
            continue
        seen_project_ids.add(pid)

        excel_row = int(i + excel_row_offset)

        raw_est = report_df.at[i, estimate_col]
        raw_x = report_df.at[i, x_col]
        raw_aa = report_df.at[i, aa_col]
        raw_ad = report_df.at[i, ad_col]

        est = _to_number(raw_est)
        vx = _to_number(raw_x) or 0.0
        vaa = _to_number(raw_aa) or 0.0
        vad = _to_number(raw_ad) or 0.0
        ssum = vx + vaa + vad

        # Pretty strings for export
        expected_str = f'אומדן פרויקט >= סה"כ עלויות צפויות {int(round(ssum))}'

        if est is None:
            actual_str = "אומדן פרויקט ריק"
            ok = False
            fail_msg = "נכשל: אומדן פרויקט ללא ערך"
        else:
            actual_str = f"אומדן פרויקט : {int(round(est))}"
            ok = est >= ssum
            fail_msg = f"נכשל: אומדן פרויקט חייב להיות >= סה\"כ עלויות צפויות {int(round(ssum))}"




        status = Status.PASS_ if ok else Status.FAIL
        sev = Severity.INFO if ok else Severity.WARNING

        excel_cells = None
        if not ok:
            refs = []
            for c in [estimate_col, x_col, aa_col, ad_col]:
                r = _cell_ref(i, c)
                if r:
                    refs.append(r)
            excel_cells = refs or None

        results.append(
            CheckResult(
                rule_id=RULE_ID,
                rule_name=RULE_NAME,
                severity=sev,
                sheet_name=SHEET,
                status=status,
                row_index=int(i),
                column_name=estimate_key,
                key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                actual_value=actual_str,
                expected_value=expected_str,
                confidence=1.0,
                method="Compare",
                excel_cells=excel_cells,
                message=(
                    f"עבר: אומדן={int(round(est))} >= סכום={int(round(ssum))}"
                    if ok
                    else fail_msg
                ),

            )
        )

    return results


def check_020_project_status_planning_report(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    R_20 — סטטוס פרויקט (בדווח תכנון)

    Per unique 'מס' פרויקט' row:
      Excel column AH (סטטוס פרויקט) must be non-empty.
    """
    from openpyxl.utils import get_column_letter, column_index_from_string

    RULE_ID = "R_20"
    RULE_NAME = "סטטוס פרויקט (בדווח תכנון)"
    SHEET = getattr(cfg, "report_sheet_name", "גיליון דיווח")

    cols = list(report_df.columns)

    def _is_empty(v: object) -> bool:
        # If pandas returns a Series (e.g., duplicate columns), treat empty if ALL elements empty
        if isinstance(v, pd.Series):
            if v.empty:
                return True
            return all(_is_empty(x) for x in v.tolist())

        # If it's a list/tuple/ndarray (rare but possible), same logic
        if isinstance(v, (list, tuple)):
            if len(v) == 0:
                return True
            return all(_is_empty(x) for x in v)

        if v is None:
            return True

        # pd.isna on scalars only
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass

        if isinstance(v, str) and v.strip() == "":
            return True

        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none"}


    # Excel -> df column by letter (A=1)
    def _col_by_excel_letter(letter: str) -> Optional[str]:
        idx_1b = column_index_from_string(letter)  # A=1
        idx_0b = idx_1b - 1
        if idx_0b < 0 or idx_0b >= len(cols):
            return None
        return cols[idx_0b]

    # Robust header normalization (same spirit as R_19)
    def _norm_col(c: object) -> str:
        s = str(c) if c is not None else ""
        s = s.replace("\n", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    norm_to_orig = {_norm_col(c): c for c in cols}

    def _resolve_by_prefix(key: str) -> Optional[str]:
        if key in norm_to_orig:
            return norm_to_orig[key]
        for norm, orig in norm_to_orig.items():
            if norm.startswith(key):
                return orig
        for norm, orig in norm_to_orig.items():
            if key in norm:
                return orig
        return None

    # Required columns
    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")
    id_col = _resolve_by_prefix(id_norm)

    status_col = _col_by_excel_letter("AF")  # <-- Excel letter AH

    # OPTIONAL debug print (like R18/R19)
    print("R20 resolved:", {"id": id_col, "AF": status_col})

    missing = []
    if id_col is None:
        missing.append(id_norm)
    if status_col is None:
        missing.append("AF")

    if missing:
        return [
            CheckResult(
                rule_id=f"{RULE_ID}_מבנה עמודות",
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name=None,
                key_context="columns_presence",
                actual_value=list(report_df.columns),
                expected_value=[id_norm, "AH"],
                message=f"Missing required columns: {missing}",
            )
        ]

    # Excel highlighting mapping
    col_to_excel_letter = {col: get_column_letter(i + 1) for i, col in enumerate(cols)}
    header_row = getattr(cfg, "report_header_row", 6)
    excel_row_offset = header_row + 3  # same as R12/R18/R19

    def _cell_ref(df_i: int, col_name: str) -> Optional[str]:
        letter = col_to_excel_letter.get(col_name)
        if not letter:
            return None
        excel_row = int(df_i + excel_row_offset)
        return f"{SHEET}!{letter}{excel_row}"

    def _is_real_row(df_i: int) -> bool:
        v = report_df.at[df_i, id_col]
        if _is_empty(v):
            return False
        s = str(v).strip()
        if s in {"-", ""}:
            return False
        if s.lower() == "nan":
            return False
        return True

    results: List[CheckResult] = []
    seen_project_ids: set[str] = set()

    for i in range(len(report_df)):
        if not _is_real_row(i):
            continue

        pid = str(report_df.at[i, id_col]).strip()
        if pid in seen_project_ids:
            continue
        seen_project_ids.add(pid)

        excel_row = int(i + excel_row_offset)

        raw_status = report_df.at[i, status_col]
        ok = not _is_empty(raw_status)

        status = Status.PASS_ if ok else Status.FAIL
        sev = Severity.INFO if ok else Severity.CRITICAL

        if ok:
            actual_str = f"סטטוס : {str(raw_status).strip()}"
            msg = "עבר: סטטוס קיים"
            excel_cells = None
        else:
            actual_str = "סטטוס ריק"
            msg = "נכשל: סטטוס פרויקט ללא ערך"
            ref = _cell_ref(i, status_col)
            excel_cells = [ref] if ref else None

        results.append(
            CheckResult(
                rule_id=RULE_ID,
                rule_name=RULE_NAME,
                severity=sev,
                sheet_name=SHEET,
                status=status,
                row_index=int(i),
                column_name="סטטוס פרויקט",
                key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                actual_value=actual_str,
                expected_value="מכיל ערך",
                confidence=1.0,
                method="NotEmpty",
                excel_cells=excel_cells,
                message=msg,
            )
        )

    return results


# =============================================================================
# R_21 — קפיצת קוטר – בדיקת שורה תואמת (Step 1)
# =============================================================================

# Authoritative diameter table (ordered ascending)
_GRADE_INCH = [4, 6, 8, 10, 12, 14, 16, 18, 20, 24, 28, 30, 32, 36, 40, 42, 46, 48, 54, 60, 64]
_GRADE_MM = [100, 160, 200, 250, 315, 355, 400, 450, 500, 630, 710, 750, 800, 914, 1000, 1050, 1150, 1200, 1350, 1500, 1600]

_GRADE_INCH_IDX = {v: i for i, v in enumerate(_GRADE_INCH)}
_GRADE_MM_IDX = {v: i for i, v in enumerate(_GRADE_MM)}


def _normalize_to_nearest(value: int, allowed: list) -> int:
    """Snap *value* to the closest entry in *allowed*.

    Tie-break (equal distance to two neighbours): choose the **larger** value.
    Examples (mm list):  225 → 250,  224 → 200,  226 → 250.
    """
    best = allowed[0]
    best_dist = abs(value - best)
    for v in allowed[1:]:
        d = abs(value - v)
        if d < best_dist or (d == best_dist and v > best):
            best = v
            best_dist = d
    return best


def check_021_diameter_jump_matching_row(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    R_21 — קפיצת קוטר: בדיקת שורה תואמת + בדיקת יחס עלות

    For rows where F == "שיקום ושדרוג":
      Step 1: Parse colon-separated diameters in I (old) and L (new),
              detect unit. If any pair has L > I by more than 2" (inch)
              or 50mm, search for a matching row with same C,D,E and
              F in ("פיתוח", "שדרוג"). FAIL if no match.
      Step 2: When matching row found, verify cost split (AE column)
              matches expected split from diameter ratios.
              rehab_share = (old_d/new_d)^1.2, weighted by segment lengths.
              PASS if within 5% tolerance.
    """
    from openpyxl.utils import get_column_letter, column_index_from_string

    RULE_ID = "R_21"
    RULE_NAME = "קפיצת קוטר – בדיקת שורה תואמת + יחס עלות"
    SHEET = getattr(cfg, "report_sheet_name", "גיליון דיווח")

    cols = list(report_df.columns)

    # ---- local helpers ----

    def _is_empty(v: object) -> bool:
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        if isinstance(v, str) and v.strip() == "":
            return True
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none"}

    def _to_number(v: object) -> Optional[float]:
        if _is_empty(v):
            return None
        try:
            s = str(v).strip().replace(",", "").replace('"', '')
            return float(s)
        except Exception:
            return None

    def _col_by_excel_letter(letter: str) -> Optional[str]:
        idx_1b = column_index_from_string(letter)
        idx_0b = idx_1b - 1
        if idx_0b < 0 or idx_0b >= len(cols):
            return None
        return cols[idx_0b]

    def _norm_col(c: object) -> str:
        s = str(c) if c is not None else ""
        s = re.sub(r"\s*Unnamed:.*", "", s)
        s = re.sub(r"_level_\d+", "", s)
        s = s.replace("\n", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    # ---- resolve columns ----
    col_c = _col_by_excel_letter("C")   # ישוב
    col_d = _col_by_excel_letter("D")   # שם פרויקט
    col_e = _col_by_excel_letter("E")   # רחוב/שכונה
    col_f = _col_by_excel_letter("F")   # סיווג פרויקט
    col_i = _col_by_excel_letter("I")   # קוטר קיים
    col_j = _col_by_excel_letter("J")   # אורך לפי קוטר קיים (colon-separated)
    col_l = _col_by_excel_letter("L")   # קוטר מתוכנן
    col_m = _col_by_excel_letter("M")   # אורך לפי קוטר מתוכנן (colon-separated)
    col_ae = _col_by_excel_letter("AE") # אומדן פרויקט

    missing = []
    for ltr, c in [("C", col_c), ("D", col_d), ("E", col_e),
                    ("F", col_f), ("I", col_i), ("L", col_l)]:
        if c is None:
            missing.append(ltr)

    if missing:
        return [
            CheckResult(
                rule_id=f"{RULE_ID}_מבנה עמודות",
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name=None,
                key_context="columns_presence",
                actual_value=list(report_df.columns),
                expected_value=["C", "D", "E", "F", "I", "L"],
                message=f"Missing required columns: {missing}",
            )
        ]

    # ---- project-id for real-row detection ----
    norm_to_orig: dict = {}
    for c in cols:
        n = _norm_col(c)
        if n and n not in norm_to_orig:
            norm_to_orig[n] = c

    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")
    id_col = norm_to_orig.get(id_norm)

    def _is_real_row(df_i: int) -> bool:
        if id_col is None:
            return True
        v = report_df.at[df_i, id_col]
        if _is_empty(v):
            return False
        s = str(v).strip()
        return s not in {"-", ""} and s.lower() != "nan"

    # ---- excel cell refs ----
    col_to_excel_letter = {col: get_column_letter(i + 1) for i, col in enumerate(cols)}
    header_row = getattr(cfg, "report_header_row", 6)
    excel_row_offset = header_row + 3

    def _cell_ref(df_i: int, col_name: str) -> Optional[str]:
        letter = col_to_excel_letter.get(col_name)
        if not letter:
            return None
        return f"{SHEET}!{letter}{int(df_i + excel_row_offset)}"

    # ---- diameter parsing helpers ----

    def _parse_colon_list(raw) -> List[str]:
        """Split colon-separated diameter string into tokens."""
        s = str(raw).strip()
        tokens = [t.strip() for t in s.split(":") if t.strip()]
        return tokens

    def _detect_unit(tokens: List[str]) -> Optional[str]:
        """Detect unit: 'inch' or 'mm'. Returns None if ambiguous."""
        has_quote = any('"' in str(t) for t in tokens)
        if has_quote:
            return "inch"

        nums = []
        for t in tokens:
            n = _to_number(t)
            if n is not None and n > 0:
                nums.append(n)

        if not nums:
            return None

        max_val = max(nums)
        if max_val <= 64:
            return "inch"
        if max_val >= 80:
            return "mm"
        return None  # ambiguous

    def _grade_index(val_num: float, unit: str) -> Optional[int]:
        """Look up grade index for a numeric diameter value (after normalization)."""
        allowed = _GRADE_INCH if unit == "inch" else _GRADE_MM
        idx_map = _GRADE_INCH_IDX if unit == "inch" else _GRADE_MM_IDX
        norm = _normalize_to_nearest(int(round(val_num)), allowed)
        return idx_map.get(norm)

    def _normalize_diam(val_num: float, unit: str) -> int:
        """Normalize a raw diameter to the nearest standard value."""
        allowed = _GRADE_INCH if unit == "inch" else _GRADE_MM
        return _normalize_to_nearest(int(round(val_num)), allowed)

    def _normalize_for_match(v) -> str:
        """Normalize a cell value for row matching (C, D, E)."""
        s = str(v).strip() if not _is_empty(v) else ""
        s = re.sub(r"\s+", " ", s)
        return s.lower()

    # ---- Step 2 helpers ----

    def _parse_num_list(raw, sep: str = ":") -> List[Optional[float]]:
        """Parse colon-separated numeric list. Returns list of floats (None for unparseable)."""
        if _is_empty(raw):
            return []
        tokens = [t.strip() for t in str(raw).strip().split(sep) if t.strip()]
        result = []
        for t in tokens:
            n = _to_number(t)
            result.append(n)
        return result

    def _parse_money_ae(raw) -> Optional[float]:
        """Parse AE monetary value (may contain commas, currency symbols, etc.)."""
        if _is_empty(raw):
            return None
        s = str(raw).strip()
        s = re.sub(r"[₪,$€\s]", "", s)
        s = s.replace(",", "")
        try:
            v = float(s)
            return v if v > 0 else None
        except (ValueError, TypeError):
            return None

    # Predefined rehab ratio table: (old_d, new_d) -> rehab share.
    # Business-defined values; do NOT derive via formula.
    # Stub: covers common inch jumps. Extend as authoritative data arrives.
    _REHAB_RATIO_TABLE: Dict[Tuple[int, int], float] = {
        (4, 8): 0.45,
        (4, 10): 0.35,
        (6, 10): 0.55,
        (6, 12): 0.45,
        (8, 12): 0.60,
        (8, 14): 0.50,
        (8, 16): 0.43,
        (10, 14): 0.65,
        (10, 16): 0.55,
        (10, 18): 0.48,
        (12, 16): 0.70,
        (12, 18): 0.61,
        (12, 20): 0.53,
        (12, 24): 0.43,
        (14, 20): 0.64,
        (14, 24): 0.51,
        (16, 24): 0.60,
        (16, 28): 0.50,
        (18, 24): 0.70,
        (18, 28): 0.57,
        (18, 30): 0.53,
        (20, 28): 0.65,
        (20, 30): 0.60,
        (24, 32): 0.70,
        (24, 36): 0.60,
        (28, 36): 0.72,
        (30, 40): 0.68,
        (32, 42): 0.70,
        (36, 48): 0.68,
        (40, 54): 0.68,
        # mm equivalents (common)
        (100, 200): 0.43,
        (100, 250): 0.35,
        (160, 250): 0.58,
        (160, 315): 0.45,
        (200, 315): 0.58,
        (200, 355): 0.50,
        (200, 400): 0.43,
        (250, 355): 0.65,
        (250, 400): 0.55,
        (250, 450): 0.48,
        (300, 400): 0.70,
        (315, 450): 0.63,
        (315, 500): 0.55,
        (355, 500): 0.64,
        (400, 630): 0.55,
        (400, 500): 0.75,
        (450, 630): 0.64,
        (500, 710): 0.63,
        (500, 630): 0.75,
        (630, 800): 0.75,
        (630, 1000): 0.55,
    }

    def _get_rehab_ratio(old_d: float, new_d: float) -> Optional[float]:
        """Look up predefined rehab ratio for (old_d -> new_d). Returns None if not in table."""
        key = (int(round(old_d)), int(round(new_d)))
        return _REHAB_RATIO_TABLE.get(key)

    STEP2_TOLERANCE = 0.05  # 5 percentage points

    # ---- main iteration ----
    results: List[CheckResult] = []

    for i in range(len(report_df)):
        if not _is_real_row(i):
            continue

        excel_row = int(i + excel_row_offset)
        pid = str(report_df.at[i, id_col]).strip() if id_col else f"row_{i}"

        # Gate: F must be "שיקום ושדרוג" or "שיקום/שדרוג"
        raw_f = str(report_df.at[i, col_f] or "").strip()
        f_norm = normalize_text(raw_f)
        if f_norm not in (normalize_text("שיקום ושדרוג"), normalize_text("שיקום/שדרוג")):
            continue

        # Gate: I and L must be non-empty
        raw_i = report_df.at[i, col_i]
        raw_l = report_df.at[i, col_l]
        if _is_empty(raw_i) or _is_empty(raw_l):
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    row_index=int(i),
                    column_name="I/L",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=f"I={'(ריק)' if _is_empty(raw_i) else str(raw_i).strip()}, L={'(ריק)' if _is_empty(raw_l) else str(raw_l).strip()}",
                    expected_value="I and L non-empty",
                    message=f"דילוג: חסר קוטר בעמודות I/L | F={raw_f}",
                    excel_cells=[c for c in [_cell_ref(i, col_i), _cell_ref(i, col_l)] if c],
                )
            )
            continue

        # Parse colon-separated lists
        tokens_i = _parse_colon_list(raw_i)
        tokens_l = _parse_colon_list(raw_l)

        # Same length check
        if len(tokens_i) != len(tokens_l):
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    row_index=int(i),
                    column_name="I/L",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=f"I={len(tokens_i)} tokens, L={len(tokens_l)} tokens",
                    expected_value="same number of tokens",
                    message=f"דילוג: אורך רשימות קוטר שונה (I:{len(tokens_i)}, L:{len(tokens_l)})",
                    excel_cells=[c for c in [_cell_ref(i, col_i), _cell_ref(i, col_l)] if c],
                )
            )
            continue

        # Detect unit
        all_tokens = tokens_i + tokens_l
        unit = _detect_unit(all_tokens)
        if unit is None:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=int(i),
                    column_name="I/L",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=f"I={str(raw_i).strip()}, L={str(raw_l).strip()}",
                    expected_value="unit detection (inch or mm)",
                    message="נכשל: לא ניתן לזהות יחידות קוטר (אינץ' או מ\"מ)",
                    excel_cells=[c for c in [_cell_ref(i, col_i), _cell_ref(i, col_l)] if c],
                )
            )
            continue

        grade_table = _GRADE_INCH_IDX if unit == "inch" else _GRADE_MM_IDX

        # Check each pair for jumps > 1 grade (using normalized diameters)
        jumps = []       # (idx, norm_old, norm_new, delta)
        norm_notes = []  # normalization annotations
        has_unrecognized = False
        for j in range(len(tokens_i)):
            num_i = _to_number(tokens_i[j])
            num_l = _to_number(tokens_l[j])
            if num_i is None or num_l is None:
                has_unrecognized = True
                continue

            raw_i_int = int(round(num_i))
            raw_l_int = int(round(num_l))
            norm_i = _normalize_diam(num_i, unit)
            norm_l = _normalize_diam(num_l, unit)

            if raw_i_int != norm_i:
                norm_notes.append(f"I[{j}]: {raw_i_int}→{norm_i}")
            if raw_l_int != norm_l:
                norm_notes.append(f"L[{j}]: {raw_l_int}→{norm_l}")

            gi = _grade_index(num_i, unit)
            gl = _grade_index(num_l, unit)

            if gi is None or gl is None:
                has_unrecognized = True
                continue

            # Regulator rule: trigger if L > I by more than 2" (inch) or 50mm
            diff = num_l - num_i
            if (unit == "inch" and diff > 2.0) or (unit == "mm" and diff > 50.0):
                jumps.append((j, norm_i, norm_l, diff))

        if has_unrecognized and not jumps:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=int(i),
                    column_name="I/L",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=f"I={str(raw_i).strip()}, L={str(raw_l).strip()}",
                    expected_value=f"recognized {unit} values in grade table",
                    message=f"נכשל: קוטר לא סטנדרטי בטבלת מדרגות ({unit}) | I_list={tokens_i} L_list={tokens_l}",
                    excel_cells=[c for c in [_cell_ref(i, col_i), _cell_ref(i, col_l)] if c],
                )
            )
            continue

        # Build per-pair debug info (with normalization)
        pair_details = []
        for j in range(len(tokens_i)):
            ni = _to_number(tokens_i[j])
            nl = _to_number(tokens_l[j])
            if ni is not None:
                ni_norm = _normalize_diam(ni, unit)
                ni_lbl = f"{int(round(ni))}({ni_norm})" if int(round(ni)) != ni_norm else str(ni_norm)
            else:
                ni_lbl = str(tokens_i[j])
                ni_norm = None
            if nl is not None:
                nl_norm = _normalize_diam(nl, unit)
                nl_lbl = f"{int(round(nl))}({nl_norm})" if int(round(nl)) != nl_norm else str(nl_norm)
            else:
                nl_lbl = str(tokens_l[j])
                nl_norm = None
            gi_dbg = _grade_index(ni, unit) if ni else None
            gl_dbg = _grade_index(nl, unit) if nl else None
            diff_dbg = round(nl - ni, 2) if (ni is not None and nl is not None) else "?"
            pair_details.append(f"{ni_lbl}[{gi_dbg}]→{nl_lbl}[{gl_dbg}] diff={diff_dbg}")
        norm_str = f" | normalized: {', '.join(norm_notes)}" if norm_notes else ""
        pairs_str = "; ".join(pair_details)

        # No significant jumps → emit debug result
        if not jumps:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    row_index=int(i),
                    column_name="I/L",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=f"I={str(raw_i).strip()}, L={str(raw_l).strip()} ({unit})",
                    expected_value="jump > 2\" (or > 50mm)",
                    message=f"דילוג: אין קפיצת קוטר מעל 2\" / 50mm | pairs: {pairs_str}{norm_str}",
                    excel_cells=[c for c in [_cell_ref(i, col_i), _cell_ref(i, col_l)] if c],
                )
            )
            continue

        # ---- Step 1: Find matching row ----
        # Normalize C, D, E of current row for comparison
        c_val = _normalize_for_match(report_df.at[i, col_c])
        d_val = _normalize_for_match(report_df.at[i, col_d])
        e_val = _normalize_for_match(report_df.at[i, col_e])

        match_found = False
        match_row_excel = None

        for r2 in range(len(report_df)):
            if r2 == i:
                continue
            if not _is_real_row(r2):
                continue

            # F must be "פיתוח" or "שדרוג"
            r2_f = normalize_text(str(report_df.at[r2, col_f] or "").strip())
            if r2_f not in (normalize_text("פיתוח"), normalize_text("שדרוג")):
                continue

            # C, D, E must match
            r2_c = _normalize_for_match(report_df.at[r2, col_c])
            r2_d = _normalize_for_match(report_df.at[r2, col_d])
            r2_e = _normalize_for_match(report_df.at[r2, col_e])

            if r2_c == c_val and r2_d == d_val and r2_e == e_val:
                match_found = True
                match_row_excel = int(r2 + excel_row_offset)
                break

        jump_desc = ", ".join(f"{ji[1]}→{ji[2]} (diff={round(ji[3], 2)})" for ji in jumps)
        sys_note = f"I_list={tokens_i} L_list={tokens_l} unit={unit} jumps=[{jump_desc}]{norm_str}"

        # Also capture the matched row's F classification and df index
        match_class = None
        match_df_idx = None
        if match_found:
            match_df_idx = match_row_excel - excel_row_offset
            r2_f_raw = str(report_df.at[match_df_idx, col_f] or "").strip()
            match_class = r2_f_raw

        if not match_found:
            # Step 1 FAIL: no matching dev row found
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.CRITICAL,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=int(i),
                    column_name="קוטר I→L",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=f"I={str(raw_i).strip()}, L={str(raw_l).strip()} ({unit})",
                    expected_value="שורת פיתוח/שדרוג תואמת",
                    confidence=1.0,
                    method="DiameterJumpMatch",
                    excel_cells=[c for c in [_cell_ref(i, col_f), _cell_ref(i, col_i), _cell_ref(i, col_l)] if c],
                    message=(
                        f"נכשל: קפיצת קוטר ({jump_desc}) – "
                        f"נדרשת שורת פיתוח/שדרוג תואמת אך לא נמצאה "
                        f"(C={c_val} D={d_val} E={e_val}) | {sys_note}"
                    ),
                )
            )
            continue

        # ================================================================
        # Step 2: Ratio check — compare expected vs actual AE split
        # ================================================================

        # --- Parse AE values ---
        rehab_cost = _parse_money_ae(report_df.at[i, col_ae]) if col_ae else None
        dev_cost = _parse_money_ae(report_df.at[match_df_idx, col_ae]) if col_ae else None

        if rehab_cost is None or dev_cost is None:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=int(i),
                    column_name="AE",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=f"rehab_AE={report_df.at[i, col_ae] if col_ae else '?'}, dev_AE={report_df.at[match_df_idx, col_ae] if col_ae else '?'}",
                    expected_value="AE > 0 for both rows",
                    confidence=1.0,
                    method="DiameterJumpRatio",
                    excel_cells=[c for c in [_cell_ref(i, col_ae), _cell_ref(match_df_idx, col_ae)] if c and col_ae],
                    message=(
                        f"נכשל: אין אומדן תקין (AE) לבדיקת יחס | "
                        f"rehab_row={excel_row} dev_row={match_row_excel} | {sys_note}"
                    ),
                )
            )
            continue

        total_cost = rehab_cost + dev_cost
        if total_cost <= 0:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=int(i),
                    column_name="AE",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=f"rehab_cost={rehab_cost}, dev_cost={dev_cost}, total=0",
                    expected_value="total AE > 0",
                    confidence=1.0,
                    method="DiameterJumpRatio",
                    message=f"נכשל: אומדן כולל = 0 | {sys_note}",
                )
            )
            continue

        rehab_share_actual = rehab_cost / total_cost

        # --- Build segments for weighted average ---
        # Try dev row diameters + lengths first
        dev_diams = _parse_num_list(report_df.at[match_df_idx, col_l]) if col_l else []
        dev_lens = _parse_num_list(report_df.at[match_df_idx, col_m]) if col_m else []

        # Rehab row L_list as normalized values for lookup
        rehab_l_nums = [_to_number(t) for t in tokens_l]
        rehab_l_norms = [_normalize_diam(v, unit) if v is not None else None for v in rehab_l_nums]
        rehab_i_nums = [_to_number(t) for t in tokens_i]

        segments = []  # list of (norm_old, norm_new, weight, ratio)
        segment_method = ""
        segment_notes = []

        # Normalize dev diameters too
        dev_diams_norm = [_normalize_diam(d, unit) if d is not None else None for d in dev_diams] if dev_diams else []

        if dev_diams_norm and all(d is not None for d in dev_diams_norm):
            # Strategy A: use dev row's diameter list mapped into rehab L_list
            segment_method = "dev_row"
            dev_lens_valid = (
                dev_lens
                and len(dev_lens) == len(dev_diams_norm)
                and all(ln is not None and ln > 0 for ln in dev_lens)
            )
            if not dev_lens_valid and dev_lens:
                segment_notes.append(f"dev_lens count mismatch or invalid ({len(dev_lens)} vs {len(dev_diams_norm)}), using equal weights")

            for k, norm_new_d in enumerate(dev_diams_norm):
                if norm_new_d is None or norm_new_d <= 0:
                    continue
                # Find matching position in rehab L_list (normalized)
                matched_idx = None
                for ridx, rl_norm in enumerate(rehab_l_norms):
                    if rl_norm is not None and rl_norm == norm_new_d:
                        matched_idx = ridx
                        break
                if matched_idx is not None:
                    old_raw = rehab_i_nums[matched_idx]
                    if old_raw is not None and old_raw > 0:
                        norm_old_d = _normalize_diam(old_raw, unit)
                        ratio = _get_rehab_ratio(norm_old_d, norm_new_d)
                        if ratio is not None:
                            w = dev_lens[k] if dev_lens_valid else 1.0
                            segments.append((norm_old_d, norm_new_d, w, ratio))
                            segment_notes.append(f"dev[{k}]: {norm_old_d}→{norm_new_d} ratio={ratio} w={w}")
                        else:
                            segment_notes.append(f"dev[{k}]: {norm_old_d}→{norm_new_d} NOT IN RATIO TABLE")
                    else:
                        segment_notes.append(f"dev[{k}]: new_d={norm_new_d} matched rehab L[{matched_idx}] but old_d invalid")
                else:
                    raw_d = dev_diams[k] if dev_diams else '?'
                    segment_notes.append(f"dev[{k}]: new_d={norm_new_d} (raw={raw_d}) not found in rehab L_list")

        if not segments:
            # Strategy B: fall back to rehab row's significant jumps from Step 1
            segment_method = "rehab_jumps" if not segment_method else f"{segment_method}→rehab_jumps"
            # Parse rehab lengths from J (for I diameters) or M (for L diameters)
            rehab_lens_j = _parse_num_list(report_df.at[i, col_j]) if col_j else []
            rehab_lens_m = _parse_num_list(report_df.at[i, col_m]) if col_m else []
            # Prefer M lengths (aligned with L diameters)
            rehab_lens = rehab_lens_m if rehab_lens_m else rehab_lens_j
            lens_valid = (
                rehab_lens
                and len(rehab_lens) == len(tokens_l)
                and all(ln is not None and ln > 0 for ln in rehab_lens)
            )
            if rehab_lens and not lens_valid:
                segment_notes.append(f"rehab lens count mismatch ({len(rehab_lens)} vs {len(tokens_l)}), using equal weights")

            for (j_idx, old_d_int, new_d_int, _delta) in jumps:
                ratio = _get_rehab_ratio(float(old_d_int), float(new_d_int))
                if ratio is not None:
                    w = rehab_lens[j_idx] if lens_valid else 1.0
                    segments.append((float(old_d_int), float(new_d_int), w, ratio))
                    segment_notes.append(f"jump[{j_idx}]: {old_d_int}→{new_d_int} ratio={ratio} w={w}")
                else:
                    segment_notes.append(f"jump[{j_idx}]: {old_d_int}→{new_d_int} NOT IN RATIO TABLE")

        # --- Compute weighted average rehab share ---
        if not segments:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    row_index=int(i),
                    column_name="קוטר I→L",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=f"no valid segments for ratio calc",
                    expected_value="at least one diameter segment",
                    method="DiameterJumpRatio",
                    message=(
                        f"דילוג: לא נמצאו מקטעים בטבלת היחסים לחישוב | "
                        f"segments_tried=[{'; '.join(segment_notes)}] | "
                        f"dev_diams={dev_diams} dev_lens={dev_lens} | {sys_note}"
                    ),
                )
            )
            continue

        total_weight = sum(s[2] for s in segments)
        if total_weight <= 0:
            total_weight = len(segments)
            segments = [(s[0], s[1], 1.0, s[3]) for s in segments]

        # Weighted average using predefined ratios from table
        rehab_share_expected = sum(
            w * ratio for _od, _nd, w, ratio in segments
        ) / total_weight

        # --- Compare ---
        diff = abs(rehab_share_actual - rehab_share_expected)
        ratio_ok = diff <= STEP2_TOLERANCE

        # --- Build system note ---
        step2_note = (
            f"key=({c_val}|{d_val}|{e_val}) "
            f"rehab_row={excel_row} dev_row={match_row_excel} dev_class={match_class} | "
            f"rehab_I={tokens_i} rehab_L={tokens_l} "
            f"dev_diams={[d for d in dev_diams]} dev_lens={[l for l in dev_lens]} | "
            f"segments=[{'; '.join(segment_notes)}] method={segment_method} | "
            f"rehab_share_expected={rehab_share_expected:.4f} "
            f"rehab_share_actual={rehab_share_actual:.4f} diff={diff:.4f} "
            f"tolerance={STEP2_TOLERANCE} | "
            f"rehab_cost={rehab_cost} dev_cost={dev_cost} total={total_cost}"
        )

        if ratio_ok:
            status = Status.PASS_
            sev = Severity.INFO
            msg = (
                f"עבר: קפיצת קוטר ({jump_desc}) – "
                f"שורת {match_class} תואמת בשורה {match_row_excel}, "
                f"יחס שיקום בפועל {rehab_share_actual:.1%} "
                f"קרוב לצפוי {rehab_share_expected:.1%} (Δ{diff:.1%}) | {step2_note}"
            )
            excel_cells = None
        else:
            status = Status.FAIL
            sev = Severity.CRITICAL
            msg = (
                f"נכשל: יחס לא תקין – קפיצת קוטר ({jump_desc}) – "
                f"שורת {match_class} בשורה {match_row_excel}: "
                f"יחס שיקום בפועל {rehab_share_actual:.1%} "
                f"שונה מהצפוי {rehab_share_expected:.1%} (Δ{diff:.1%}, "
                f"סף={STEP2_TOLERANCE:.0%}) | {step2_note}"
            )
            excel_cells = [c for c in [
                _cell_ref(i, col_ae), _cell_ref(match_df_idx, col_ae),
                _cell_ref(i, col_i), _cell_ref(i, col_l),
            ] if c]

        results.append(
            CheckResult(
                rule_id=RULE_ID,
                rule_name=RULE_NAME,
                severity=sev,
                sheet_name=SHEET,
                status=status,
                row_index=int(i),
                column_name="קוטר I→L / AE",
                key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                actual_value=f"rehab_share={rehab_share_actual:.1%} (rehab={rehab_cost}, dev={dev_cost})",
                expected_value=f"rehab_share≈{rehab_share_expected:.1%} ±{STEP2_TOLERANCE:.0%}",
                confidence=1.0,
                method="DiameterJumpRatio",
                excel_cells=excel_cells,
                message=msg,
            )
        )

    return results


# =============================================================================
# R_23 — בדיקת מחיר צנרת - לפי כלל אצבע
# =============================================================================

# mm → inch conversion built from R_21 authoritative tables
_MM_TO_INCH: Dict[int, int] = {mm: inch for mm, inch in zip(_GRADE_MM, _GRADE_INCH)}

# Values at or below this threshold are treated as inches; above as mm
_INCH_THRESHOLD = 64


def check_023_pipe_cost_rule_of_thumb(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    R_23 — בדיקת מחיר צנרת לפי כלל אצבע (מפקח)

    Applies only to rows where column G == "קווי מים" or "קווי ביוב".
    Regulator formula:
        Total Cost = Column AE (אלפי ₪)
        Total Length = Sum of values in Column M (split by :)
        Cost per meter (₪/m) = (AE * 1000) / Total Length
        Baseline (₪/m) = 0.15 * 1000 * diameter_inches = 150 * diameter_inches
        Tolerance: ±20% → allowed [0.8*baseline, 1.2*baseline]
    Uses max diameter when L has multiple values (e.g. 4:6:8).
    """
    from openpyxl.utils import get_column_letter, column_index_from_string

    RULE_ID = "R_23"
    RULE_NAME = "בדיקת מחיר צנרת - לפי כלל אצבע"
    SHEET = getattr(cfg, "report_sheet_name", "גיליון דיווח")

    cols = list(report_df.columns)

    # ---- local helpers (same pattern as R19/R24) ----

    def _is_empty(v: object) -> bool:
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        if isinstance(v, str) and v.strip() == "":
            return True
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none"}

    def _to_number(v: object) -> Optional[float]:
        if _is_empty(v):
            return None
        try:
            s = str(v).strip().replace(",", "")
            return float(s)
        except Exception:
            return None

    def _col_by_excel_letter(letter: str) -> Optional[str]:
        idx_1b = column_index_from_string(letter)
        idx_0b = idx_1b - 1
        if idx_0b < 0 or idx_0b >= len(cols):
            return None
        return cols[idx_0b]

    def _norm_col(c: object) -> str:
        s = str(c) if c is not None else ""
        s = re.sub(r"\s*Unnamed:.*", "", s)
        s = re.sub(r"_level_\d+", "", s)
        s = s.replace("\n", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    # ---- resolve columns ----
    col_g = _col_by_excel_letter("G")   # קוד הנדסי
    col_l = _col_by_excel_letter("L")   # קוטר
    col_m = _col_by_excel_letter("M")   # אורך צנרת
    col_ae = _col_by_excel_letter("AE") # אומדן פרויקט

    missing = []
    if col_g is None:
        missing.append("G")
    if col_l is None:
        missing.append("L")
    if col_m is None:
        missing.append("M")
    if col_ae is None:
        missing.append("AE")

    if missing:
        return [
            CheckResult(
                rule_id=f"{RULE_ID}_מבנה עמודות",
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name=None,
                key_context="columns_presence",
                actual_value=list(report_df.columns),
                expected_value=["G", "L", "M", "AE"],
                message=f"Missing required columns: {missing}",
            )
        ]

    # ---- project-id for real-row detection ----
    norm_to_orig: dict = {}
    for c in cols:
        n = _norm_col(c)
        if n and n not in norm_to_orig:
            norm_to_orig[n] = c

    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")
    id_col = norm_to_orig.get(id_norm)

    def _is_real_row(df_i: int) -> bool:
        if id_col is None:
            return True
        v = report_df.at[df_i, id_col]
        if _is_empty(v):
            return False
        s = str(v).strip()
        return s not in {"-", ""} and s.lower() != "nan"

    # ---- excel cell refs ----
    col_to_excel_letter = {col: get_column_letter(i + 1) for i, col in enumerate(cols)}
    header_row = getattr(cfg, "report_header_row", 6)
    excel_row_offset = header_row + 3

    def _cell_ref(df_i: int, col_name: str) -> Optional[str]:
        letter = col_to_excel_letter.get(col_name)
        if not letter:
            return None
        return f"{SHEET}!{letter}{int(df_i + excel_row_offset)}"

    # ---- diameter conversion helper (uses R_21 normalize_to_nearest) ----
    def _diameter_to_inch(raw_val) -> Tuple[Optional[float], str]:
        """
        Convert raw diameter value to inches using authoritative table.
        Normalizes to nearest standard value first.
        Returns (inch_value, note).
        """
        num = _to_number(raw_val)
        if num is None:
            return None, "ערך קוטר לא מספרי"
        if num <= 0:
            return None, "קוטר <= 0"

        raw_int = int(round(num))

        if raw_int <= _INCH_THRESHOLD:
            # Treat as inches — normalize to nearest standard inch
            norm = _normalize_to_nearest(raw_int, _GRADE_INCH)
            note = f"{raw_int}→{norm}\"" if raw_int != norm else ""
            return float(norm), note

        # Treat as mm — normalize then convert
        norm_mm = _normalize_to_nearest(raw_int, _GRADE_MM)
        inch = _MM_TO_INCH[norm_mm]
        if raw_int != norm_mm:
            return float(inch), f"המרה: {raw_int}mm→{norm_mm}mm→{inch}\""
        return float(inch), f"המרה: {norm_mm}mm→{inch}\""

    # ---- main iteration ----
    results: List[CheckResult] = []

    for i in range(len(report_df)):
        if not _is_real_row(i):
            continue

        excel_row = int(i + excel_row_offset)
        pid = str(report_df.at[i, id_col]).strip() if id_col else f"row_{i}"

        # 1. Filter: only "קווי מים" or "קווי ביוב" in column G
        raw_g = str(report_df.at[i, col_g] or "").strip()
        g_norm = normalize_text(raw_g)
        if "קווי מים" not in g_norm and "קווי ביוב" not in g_norm:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    row_index=int(i),
                    column_name="G",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=raw_g or "(ריק)",
                    expected_value="קווי מים / קווי ביוב",
                    message=f"דילוג: קוד הנדסי אינו קווי מים או קווי ביוב (G={raw_g or '(ריק)'})",
                    excel_cells=[_cell_ref(i, col_g)] if _cell_ref(i, col_g) else None,
                )
            )
            continue

        # 2. Column L empty → skip
        raw_l = report_df.at[i, col_l]
        if _is_empty(raw_l):
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=int(i),
                    column_name="L",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value="אין ערך",
                    expected_value="אין ערך",
                    message="נכשל: אין ערך בעמודת קוטר (L)",
                    excel_cells=[_cell_ref(i, col_l)] if _cell_ref(i, col_l) else None,
                )
            )
            continue

        raw_l_str = str(raw_l).strip()

        # 3. Parse diameter(s) and length(s) — supports single and multi-value
        l_tokens = [t.strip() for t in raw_l_str.split(":") if t.strip()]

        raw_m = report_df.at[i, col_m]
        raw_m_str = str(raw_m).strip() if not _is_empty(raw_m) else ""
        m_tokens = [t.strip() for t in raw_m_str.split(":") if t.strip()] if raw_m_str else []

        is_multi = len(l_tokens) > 1

        # For multi-value: count must match between L and M
        if is_multi and len(l_tokens) != len(m_tokens):
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=int(i),
                    column_name="L/M",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value="אי התאמה",
                    expected_value="אי התאמה",
                    confidence=1.0,
                    method="RuleOfThumb",
                    message=f"נכשל: אין התאמה בין אורכים לקטרים | L={raw_l_str} M={raw_m_str}",
                    excel_cells=[c for c in [_cell_ref(i, col_l), _cell_ref(i, col_m)] if c],
                )
            )
            continue

        # 4. Convert each diameter to inches
        inch_vals = []
        conv_notes = []
        bad_diam = False
        for t in l_tokens:
            d_inch, note = _diameter_to_inch(t)
            if d_inch is None:
                bad_diam = True
                conv_notes.append(f"{t}: {note}")
                break
            inch_vals.append(d_inch)
            if note:
                conv_notes.append(note)

        if bad_diam:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    row_index=int(i),
                    column_name="L",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=raw_l_str,
                    expected_value="קוטר תקין באינצ׳ או מ״מ",
                    message=f"דילוג: {'; '.join(conv_notes)}",
                    excel_cells=[_cell_ref(i, col_l)] if _cell_ref(i, col_l) else None,
                )
            )
            continue

        # 5. Parse lengths
        m_vals = []
        for t in (m_tokens if m_tokens else [raw_m_str] if raw_m_str else []):
            n = _to_number(t)
            if n is not None and n > 0:
                m_vals.append(n)
            else:
                m_vals.append(None)

        total_length = sum(v for v in m_vals if v is not None)
        if total_length <= 0:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    row_index=int(i),
                    column_name="M",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=str(raw_m),
                    expected_value="אורך צנרת > 0",
                    message="דילוג: אורך צנרת חסר או 0 (מניעת חילוק באפס)",
                    excel_cells=[_cell_ref(i, col_m)] if _cell_ref(i, col_m) else None,
                )
            )
            continue

        # 6. Column AE (estimate) must be numeric
        raw_ae = report_df.at[i, col_ae]
        ae_val = _to_number(raw_ae)
        if ae_val is None:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.NOT_APPLICABLE,
                    row_index=int(i),
                    column_name="AE",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=str(raw_ae),
                    expected_value="אומדן פרויקט מספרי",
                    message="דילוג: אומדן פרויקט חסר או לא מספרי",
                    excel_cells=[_cell_ref(i, col_ae)] if _cell_ref(i, col_ae) else None,
                )
            )
            continue

        # 7. Regulator formula: Cost per meter = (AE*1000)/Total Length; Baseline = 150 * diameter (inches)
        cost_per_m = (ae_val * 1000) / total_length
        max_diameter_inch = max(inch_vals)
        baseline_nis_per_m = 150 * max_diameter_inch  # 0.15 * 1000 * diameter
        tolerance = 0.20
        low_bound = (1 - tolerance) * baseline_nis_per_m
        high_bound = (1 + tolerance) * baseline_nis_per_m

        if is_multi:
            segment_details = [f"{d}\"×{m}" for d, m in zip(inch_vals, m_vals) if m is not None]
            conv_suffix = f" | segments: [{', '.join(segment_details)}] | {'; '.join(conv_notes)}" if conv_notes else f" | segments: [{', '.join(segment_details)}]"
            diam_display = f"max({', '.join(str(d) for d in inch_vals)}\")"
        else:
            conv_suffix = f" | {conv_notes[0]}" if conv_notes else ""
            diam_display = f"{max_diameter_inch}\""

        # 8. Deviation check — pass: cost_per_m in [0.8*baseline, 1.2*baseline]
        is_fail = cost_per_m < low_bound or cost_per_m > high_bound

        refs = []
        for c in [col_l, col_m, col_ae]:
            r = _cell_ref(i, c)
            if r:
                refs.append(r)

        expected_range = f"{int(round(low_bound))}-{int(round(high_bound))} ₪/מ' (בסיס: 150×קוטר={int(round(baseline_nis_per_m))} ₪/מ', ±20%)"
        actual_display = {"cost_per_meter_nis": round(cost_per_m, 2), "total_cost_ae": ae_val, "total_length_m": total_length, "diameter_inch": max_diameter_inch, "baseline_nis_per_m": round(baseline_nis_per_m, 2)}

        if is_fail:
            if cost_per_m < low_bound:
                fail_reason = f"עלות למטר ({round(cost_per_m, 1)} ₪) נמוכה מתחת לטווח (±20% מבסיס {int(round(baseline_nis_per_m))} ₪)"
            else:
                fail_reason = f"עלות למטר ({round(cost_per_m, 1)} ₪) מעל הטווח המותר (±20% מבסיס {int(round(baseline_nis_per_m))} ₪)"

            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=int(i),
                    column_name="AE",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=actual_display,
                    expected_value=expected_range,
                    confidence=1.0,
                    method="RuleOfThumb",
                    message=(
                        f"חריגה: {fail_reason} | "
                        f"קוטר={diam_display} | AE={ae_val} | M={total_length}{conv_suffix}"
                    ),
                    excel_cells=refs or None,
                )
            )
        else:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.PASS_,
                    row_index=int(i),
                    column_name="AE",
                    key_context=f"{id_norm}={pid} | excel_row={excel_row}",
                    actual_value=actual_display,
                    expected_value=expected_range,
                    confidence=1.0,
                    method="RuleOfThumb",
                    message=(
                        f"תקין: עלות למטר ({round(cost_per_m, 1)} ₪) "
                        f"בטווח {int(round(low_bound))}-{int(round(high_bound))} ₪/מ' | "
                        f"קוטר={diam_display}{conv_suffix}"
                    ),
                    excel_cells=refs or None,
                )
            )

    return results


def check_024_short_pipe_projects_ratio(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    R_24 — ratio of expected costs (AE) for projects whose minimum pipe length in column M is < 100.

    - Identify real project rows by project-id column (default "מס' פרויקט")
    - For each row: extract all numbers in M (supports ":" separated), take MIN.
      If MIN < 100 => row is in "small pipe" group.
    - Sum AE ONLY for rows with numeric AE (skip rows where AE is empty/unparseable)
      - If ALL rows have AE empty => CRITICAL "no info"
      - If AE has content but NONE numeric => CRITICAL "no numeric"
    - ratio = sum_small / sum_total; if ratio > 5% => FAIL (CRITICAL)
    """
    import re
    import math
    from openpyxl.utils import column_index_from_string

    RULE_ID = "R_24"
    RULE_NAME = "אחוז פרויקטים עם אורך צנרת קטן מ-100 מטרים"
    SHEET = getattr(cfg, "report_sheet_name", "גיליון דיווח")

    THRESH_M = 100.0
    THRESH_RATIO = 0.05  # 5%

    cols = list(report_df.columns)

    def _finite_or_none(x: Optional[float]) -> Optional[float]:
        try:
            if x is None:
                return None
            x = float(x)
            return x if math.isfinite(x) else None
        except Exception:
            return None

    def _safe_int(x: Optional[float]) -> int:
        x2 = _finite_or_none(x)
        return int(round(x2)) if x2 is not None else 0

    # Excel -> df column by letter (A=1)
    def _col_by_excel_letter(letter: str) -> Optional[str]:
        idx_1b = column_index_from_string(letter)
        idx_0b = idx_1b - 1
        if idx_0b < 0 or idx_0b >= len(cols):
            return None
        return cols[idx_0b]

    def _is_empty_scalar(v: object) -> bool:
        if v is None:
            return True
        if isinstance(v, (list, tuple, dict, set)):
            return len(v) == 0
        if hasattr(v, "shape") and not isinstance(v, (str, bytes)):
            return False
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        if isinstance(v, str) and v.strip() == "":
            return True
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none"}

    def _to_number(v: object) -> Optional[float]:
        if _is_empty_scalar(v):
            return None
        try:
            s = str(v).strip().replace(",", "")
            return float(s)
        except Exception:
            return None

    def _extract_numbers_min(v: object) -> Optional[float]:
        """
        Return the minimum numeric value found in the cell.
        Supports colon-separated patterns like "800:500:200".
        Falls back to regex for messy text.
        """
        if _is_empty_scalar(v):
            return None

        s = str(v).strip()

        # Fast path: ":" / commas / whitespace separated tokens
        parts = re.split(r"[:，,;\s]+", s)
        vals: list[float] = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            if re.fullmatch(r"-?\d+(?:\.\d+)?", p):
                try:
                    vals.append(float(p))
                except Exception:
                    pass
        if vals:
            return min(vals)

        # Fallback: find numbers anywhere
        nums = re.findall(r"-?\d+(?:\.\d+)?", s)
        if not nums:
            return None
        vals = []
        for t in nums:
            try:
                vals.append(float(t))
            except Exception:
                continue
        return min(vals) if vals else None

    def _extract_length_sum(v: object) -> Optional[float]:
        """Sum of numeric values in M (split by colon). Used for total length < 100m check."""
        if _is_empty_scalar(v):
            return None
        s = str(v).strip()
        parts = [t.strip() for t in s.split(":") if t.strip()]
        vals = []
        for p in parts:
            if re.fullmatch(r"-?\d+(?:\.\d+)?", p):
                try:
                    vals.append(float(p))
                except Exception:
                    pass
        if not vals:
            nums = re.findall(r"-?\d+(?:\.\d+)?", s)
            for t in nums:
                try:
                    vals.append(float(t))
                except Exception:
                    pass
        return sum(vals) if vals else None

    # Resolve project id column by normalized header
    def _norm_col(c: object) -> str:
        s = str(c) if c is not None else ""
        s = re.sub(r"\s*Unnamed:.*", "", s)
        s = re.sub(r"_level_\d+", "", s)
        s = s.replace("\n", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    norm_to_orig: dict[str, str] = {}
    for c in cols:
        n = _norm_col(c)
        if n and n not in norm_to_orig:
            norm_to_orig[n] = c

    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")
    id_col = norm_to_orig.get(id_norm)

    col_m = _col_by_excel_letter("M")
    col_ae = _col_by_excel_letter("AE")
    col_c = _col_by_excel_letter("C")  # project name for context
    header_row = getattr(cfg, "report_header_row", 6)

    print("R24 resolved:", {"id": id_col, "M": col_m, "AE": col_ae})

    missing = []
    if id_col is None:
        missing.append(id_norm)
    if col_m is None:
        missing.append("M")
    if col_ae is None:
        missing.append("AE")

    if missing:
        return [
            CheckResult(
                rule_id=f"{RULE_ID}_מבנה עמודות",
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name=None,
                key_context="columns_presence",
                actual_value=list(report_df.columns),
                expected_value=[id_norm, "M", "AE"],
                message=f"Missing required columns: {missing}",
            )
        ]

    def _is_real_row(df_i: int) -> bool:
        v = report_df.at[df_i, id_col]
        if _is_empty_scalar(v):
            return False
        s = str(v).strip()
        if s in {"-", ""}:
            return False
        if s.lower() == "nan":
            return False
        return True

    sum_total = 0.0
    sum_small = 0.0
    short_pipe_results: List[CheckResult] = []  # per-row: total length < 100m

    any_rows = 0                 # count of real project rows (by id)
    any_small = 0                # count of rows where min(M) < 100 (regardless of AE)

    ae_present_rows = 0          # AE has any content (non-empty)
    ae_numeric_rows = 0          # AE parsed to finite number
    ae_skipped_total = 0         # rows skipped from denominator
    ae_skipped_small = 0         # rows skipped from numerator (in small group)

    for i in range(len(report_df)):
        if not _is_real_row(i):
            continue
        any_rows += 1

        raw_m = report_df.at[i, col_m]
        total_length = _extract_length_sum(raw_m)
        if total_length is not None and total_length < THRESH_M:
            proj_name = report_df.at[i, col_c] if col_c is not None else report_df.at[i, id_col]
            excel_row_1based = header_row + 3 + i
            short_pipe_results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=excel_row_1based,
                    column_name=col_m,
                    key_context=str(proj_name) if proj_name is not None else "",
                    actual_value=total_length,
                    expected_value=">= 100 (מטרים)",
                    message="התרעה: קו קצר מ-100 מטר ללא הסבר נלווה",
                    confidence=1.0,
                    method="ShortPipeLength",
                )
            )
        m_min = _extract_numbers_min(raw_m)
        is_small = (m_min is not None) and (m_min < THRESH_M)
        if is_small:
            any_small += 1

        raw_ae = report_df.at[i, col_ae]
        if not _is_empty_scalar(raw_ae):
            ae_present_rows += 1

        ae_num = _to_number(raw_ae)
        if ae_num is None or (not math.isfinite(ae_num)):
            ae_skipped_total += 1
            if is_small:
                ae_skipped_small += 1
            continue

        ae_numeric_rows += 1
        sum_total += ae_num
        if is_small:
            sum_small += ae_num

    if any_rows == 0:
        return short_pipe_results + [
            CheckResult(
                rule_id=RULE_ID,
                rule_name=RULE_NAME,
                severity=Severity.WARNING,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name="M",
                key_context="no_project_rows",
                actual_value="לא נמצאו שורות פרויקט",
                expected_value="נדרש לפחות פרויקט אחד",
                confidence=1.0,
                method="Summary",
                message="נכשל: לא נמצאו שורות פרויקט לביצוע בדיקה",
            )
        ]

    # ALARM: all AE cells empty across all real project rows
    if ae_present_rows == 0:
        return short_pipe_results + [
            CheckResult(
                rule_id=RULE_ID,
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name="AE",
                key_context="ae_no_info_all_rows",
                actual_value=f"AE ריק בכל שורות הפרויקט (סה\"כ פרויקטים: {any_rows})",
                expected_value="נדרש לפחות אומדן אחד בעמודה AE",
                confidence=1.0,
                method="Summary",
                message="שגיאה קריטית: אין מידע בעמודה AE (כל הערכים ריקים), לא ניתן לחשב יחס.",
            )
        ]

    # ALARM: AE has content but none numeric/finite
    if ae_numeric_rows == 0:
        return short_pipe_results + [
            CheckResult(
                rule_id=RULE_ID,
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name="AE",
                key_context="ae_unparseable_all_rows",
                actual_value=f"AE לא מספרי בכל שורות הפרויקט (סה\"כ פרויקטים: {any_rows}, לא ריקים: {ae_present_rows})",
                expected_value="נדרש אומדן מספרי בעמודה AE",
                confidence=1.0,
                method="Summary",
                message="שגיאה קריטית: בעמודה AE אין אף ערך מספרי תקין, לא ניתן לחשב יחס.",
            )
        ]

    # After skipping missing AE rows, still must have a positive denominator
    if (not math.isfinite(sum_total)) or (sum_total <= 0):
        return short_pipe_results + [
            CheckResult(
                rule_id=RULE_ID,
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name="AE",
                key_context="total_estimate_zero_after_skips",
                actual_value=f"סה\"כ אומדנים (רק ערכים תקינים)={_safe_int(sum_total)} | דילוגים={ae_skipped_total}",
                expected_value="סה\"כ אומדנים > 0",
                confidence=1.0,
                method="Summary",
                message="נכשל: לאחר דילוג על פרויקטים ללא אומדן AE תקין, סה\"כ האומדנים הוא 0 ולכן לא ניתן לחשב יחס.",
            )
        ]

    ratio = sum_small / sum_total
    pct = ratio * 100.0

    note = ""
    if ae_skipped_total > 0:
        note = f" | דילוג על {ae_skipped_total} פרויקטים ללא אומדן AE מספרי תקין"
        if ae_skipped_small > 0:
            note += f" (מתוכם {ae_skipped_small} בקבוצת <100)"

    ok = ratio <= THRESH_RATIO
    status = Status.PASS_ if ok else Status.FAIL
    sev = Severity.INFO if ok else Severity.CRITICAL

    actual_str = f"{pct:.2f}% (סכום קטן מ-100: {_safe_int(sum_small)} מתוך {_safe_int(sum_total)}){note}"
    expected_str = "עד 5%"

    msg = (
        f"עבר: אחוז הפרויקטים עם אורך צנרת קטן מ-100 מטרים הוא {pct:.2f}%{note}"
        if ok
        else f"שגיאה: אחוז הפרויקטים עם אורך צנרת קטן מ-100 מטרים הוא {pct:.2f}% (> 5%){note}"
    )

    return short_pipe_results + [
        CheckResult(
            rule_id=RULE_ID,
            rule_name=RULE_NAME,
            severity=sev,
            sheet_name=SHEET,
            status=status,
            row_index=None,
            column_name="אורך צנרת (M) + אומדן (AE)",
            key_context=(
                f"small<100_count={any_small} | total_projects={any_rows} | "
                f"ae_numeric_rows={ae_numeric_rows} | skipped_ae_rows={ae_skipped_total}"
            ),
            actual_value=actual_str,
            expected_value=expected_str,
            confidence=1.0,
            method="SummaryRatio",
            excel_cells=None,
            message=msg,
        )
    ]


def check_025_pipe_delimiter_colon_only(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    R_25 — Data formatting (גיליון דיווח): In columns J (Existing Dia), K (Existing Length),
    L (Planned Dia), M (Planned Length), multiple values MUST be separated ONLY by colon ':'.
    If '+', '/', or '-' appear as separators, flag as Error.
    Processes from physical Excel row 9 (header_row + 3 when header is on row 6).
    """
    from openpyxl.utils import get_column_letter, column_index_from_string

    RULE_ID = "R_25"
    RULE_NAME = "מפריד עמודות צנרת - רק נקודתיים"
    SHEET = getattr(cfg, "report_sheet_name", "גיליון דיווח")
    cols = list(report_df.columns)

    # Map Excel letters J,K,L,M to dataframe column names by position (handles hidden/shifted cols)
    def _col_by_excel_letter(letter: str) -> Optional[str]:
        idx_1b = column_index_from_string(letter)
        idx_0b = idx_1b - 1
        if idx_0b < 0 or idx_0b >= len(cols):
            return None
        return cols[idx_0b]

    pipe_col_letters = ["J", "K", "L", "M"]
    pipe_cols = {letter: _col_by_excel_letter(letter) for letter in pipe_col_letters}
    missing = [letter for letter, c in pipe_cols.items() if c is None]
    if missing:
        return [
            CheckResult(
                rule_id=f"{RULE_ID}_מבנה",
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name=None,
                key_context="columns_presence",
                actual_value=list(report_df.columns),
                expected_value=pipe_col_letters,
                message=f"Missing columns: {missing}",
            )
        ]

    # Row validation helpers (same as check_021): _norm_col, id_col, _is_real_row
    def _norm_col(c: object) -> str:
        s = str(c) if c is not None else ""
        s = re.sub(r"\s*Unnamed:.*", "", s)
        s = re.sub(r"_level_\d+", "", s)
        s = s.replace("\n", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    norm_to_orig: dict = {}
    for c in cols:
        n = _norm_col(c)
        if n and n not in norm_to_orig:
            norm_to_orig[n] = c

    id_norm = getattr(cfg, "report_project_id_col_norm", "מס' פרויקט")
    id_col = norm_to_orig.get(id_norm)

    def _is_empty(v: object) -> bool:
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        if isinstance(v, str) and v.strip() == "":
            return True
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none"}

    def _is_real_row(df_i: int) -> bool:
        if id_col is None:
            return True
        v = report_df.at[df_i, id_col]
        if _is_empty(v):
            return False
        s = str(v).strip()
        return s not in {"-", ""} and s.lower() != "nan"

    # Start row: physical Excel row 9 = header_row + 3 (when header is on row 6)
    header_row = getattr(cfg, "report_header_row", 6)
    excel_row_offset = header_row + 3  # first data row in Excel (1-based)
    col_to_letter = {c: get_column_letter(i + 1) for i, c in enumerate(cols)}
    forbidden = re.compile(r"[\+\/\-]")

    def _is_empty_cell(raw: object) -> bool:
        if raw is None:
            return True
        if isinstance(raw, float):
            if pd.isna(raw) or not math.isfinite(raw):
                return True
        s = str(raw).strip()
        if not s or s.lower() in {"nan", "none"}:
            return True
        return False

    results: List[CheckResult] = []
    num_rows = len(report_df)
    for i in range(num_rows):
        if not _is_real_row(i):
            continue

        excel_row = excel_row_offset + i
        row_has_pipe_data = False
        row_errors: List[CheckResult] = []

        for letter in pipe_col_letters:
            col_name = pipe_cols[letter]
            if col_name is None:
                continue
            raw = report_df.at[i, col_name]
            if _is_empty_cell(raw):
                continue
            row_has_pipe_data = True
            s = str(raw).strip()
            if forbidden.search(s):
                cell_ref = f"{col_to_letter.get(col_name, letter)}{excel_row}"
                row_errors.append(
                    CheckResult(
                        rule_id=RULE_ID,
                        rule_name=RULE_NAME,
                        severity=Severity.WARNING,
                        sheet_name=SHEET,
                        status=Status.FAIL,
                        row_index=excel_row,
                        column_name=col_name,
                        key_context=f"cell={cell_ref}",
                        actual_value=s,
                        expected_value="ערכים מרובים מופרדים רק בנקודתיים (:)",
                        message=f"ערך מכיל מפריד לא חוקי (+ / -) בעמודה {letter}: {s[:80]}",
                    )
                )

        results.extend(row_errors)
        if row_has_pipe_data and not row_errors:
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.INFO,
                    sheet_name=SHEET,
                    status=Status.PASS_,
                    row_index=excel_row,
                    column_name=None,
                    key_context=str(report_df.at[i, id_col]).strip() if id_col else "",
                    actual_value="תקין",
                    expected_value="מפריד : בלבד",
                    message="עבר: נתוני צנרת תקינים",
                )
            )
    return results



def check_016_wells_classification(
    report_df: pd.DataFrame,
    cfg: PlanConfig,
) -> List[CheckResult]:
    """
    R_16: If project name (Column C) contains "באר" or "בארות", Column F (סיווג) MUST be
    "קידוחים" and NOT "מתקני מים".
    """
    from openpyxl.utils import get_column_letter, column_index_from_string

    RULE_ID = "R_16"
    RULE_NAME = "סיווג בארות - קידוחים"
    SHEET = getattr(cfg, "report_sheet_name", "גיליון דיווח")
    cols = list(report_df.columns)

    def _col_by_excel_letter(letter: str) -> Optional[str]:
        idx_1b = column_index_from_string(letter)
        idx_0b = idx_1b - 1
        if idx_0b < 0 or idx_0b >= len(cols):
            return None
        return cols[idx_0b]

    col_c = _col_by_excel_letter("C")
    col_f = _col_by_excel_letter("F")
    if col_c is None or col_f is None:
        return [
            CheckResult(
                rule_id=f"{RULE_ID}_מבנה",
                rule_name=RULE_NAME,
                severity=Severity.CRITICAL,
                sheet_name=SHEET,
                status=Status.FAIL,
                row_index=None,
                column_name=None,
                key_context="columns_presence",
                actual_value=list(report_df.columns),
                expected_value=["C", "F"],
                message="Missing columns C or F",
            )
        ]

    header_row = getattr(cfg, "report_header_row", 6)
    excel_row_offset = header_row + 3

    results: List[CheckResult] = []
    for i in range(len(report_df)):
        name = str(report_df.at[i, col_c] or "").strip()
        if "באר" not in name and "בארות" not in name:
            continue
        raw_f = report_df.at[i, col_f]
        f_norm = normalize_text(str(raw_f or "").strip())
        if "קידוח" in f_norm or "קידוחים" in f_norm:
            continue
        if "מתקני מים" in f_norm or "מתקן מים" in f_norm:
            excel_row = excel_row_offset + i
            results.append(
                CheckResult(
                    rule_id=RULE_ID,
                    rule_name=RULE_NAME,
                    severity=Severity.WARNING,
                    sheet_name=SHEET,
                    status=Status.FAIL,
                    row_index=excel_row,
                    column_name=col_f,
                    key_context=name,
                    actual_value=str(raw_f),
                    expected_value="קידוחים",
                    message='פרויקט המכיל "באר/בארות" חייב להיות מסווג כ"קידוחים" ולא "מתקני מים"',
                )
            )
    return results


# Regex catches things like:
# "רחוב שכונת" / "רחוב שכונה" / "רחוב שכונת ..." where there's no real descriptor
INVALID_PROJECT_REGEXES = [
    re.compile(r"^\s*רחוב\s*$"),
    re.compile(r"^\s*בין\s+הבתים\s*$"),
    re.compile(r"^\s*שטח\s+פתוח\s*$"),
    re.compile(r"^\s*רחוב\s+שכונ(?:ה|ת)\s*$"),  # "רחוב שכונה/שכונת" only
]


__all__ = [
    # ── Macro summary (R_1–R_4) ───────────────────────────────────────────────
    "check_001_kinun_values_rounded",
    "check_002_asset_ratio",
    "check_003_defined_value_percent",
    "check_004_total_program_values",
    # ── Standardised city checks (R_5–R_9) ────────────────────────────────────
    "check_005_total_planned_investments_cross_row",
    "check_006_sync_budget_sources_missing",
    "check_007_sync_budget_deficit",
    "check_008_pipe_lengths_water",
    "check_009_pipe_lengths_sewer",
    # ── Project-level checks (R_12–R_25) ──────────────────────────────────────
    "check_012_project_fields_not_empty",
    "check_014_llm_project_funding_classification",
    "check_015_invalid_project_names",
    "check_016_wells_classification",
    "check_018_facility_rehab_upgrade",
    "check_019_total_planned_cost_per_project",
    "check_020_project_status_planning_report",
    "check_021_diameter_jump_matching_row",
    "check_023_pipe_cost_rule_of_thumb",
    "check_024_short_pipe_projects_ratio",
    "check_025_pipe_delimiter_colon_only",
]
