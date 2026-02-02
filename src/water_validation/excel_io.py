# src/water_validation/io.py
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple, Optional

import pandas as pd

#from config import InputDiscoveryConfig, PlanConfig
from .config import InputDiscoveryConfig, PlanConfig

#from utils import is_city_like_header, normalize_text
from .utils import is_city_like_header, normalize_text



log = logging.getLogger(__name__)


@dataclass(frozen=True)
class PlanLoadInfo:
    data_start_idx_0based: int
    header_row_idx_0based: int
    data_start_excel_row_1based: int


def _pick_newest(paths: List[Path]) -> Path:
    if not paths:
        raise FileNotFoundError("No matching files found.")
    return max(paths, key=lambda p: p.stat().st_mtime)

def discover_inputs(input_dir: str | Path, disco: InputDiscoveryConfig) -> Tuple[List[Path], Optional[Path]]:
    """
    Find:
      - plan files: potentially many
      - one kinun file

    Looks in:
      1) input_dir
      2) input_dir / "data"   (project convention)
    """
    input_dir = Path(input_dir).expanduser().resolve()
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    # Build plan glob variants: handle both תכנית and תוכנית spellings
    plan_globs = [disco.plan_glob]
    if "תכנית" in disco.plan_glob:
        plan_globs.append(disco.plan_glob.replace("תכנית", "תוכנית"))

    # 1) search in input_dir
    plan_files_set: set[Path] = set()
    for pg in plan_globs:
        plan_files_set.update(input_dir.glob(pg))
    # Exclude temp files (e.g. ~$...)
    plan_files = sorted(
        [p for p in plan_files_set if not p.name.startswith("~$")],
        key=lambda p: p.name,
    )
    kinun_candidates = sorted(input_dir.glob(disco.kinun_glob))

    # 2) fallback: search in input_dir/data
    data_dir = input_dir / "data"
    if (not plan_files or not kinun_candidates) and data_dir.exists():
        if not plan_files:
            for pg in plan_globs:
                plan_files_set.update(data_dir.glob(pg))
            plan_files = sorted(
                [p for p in plan_files_set if not p.name.startswith("~$")],
                key=lambda p: p.name,
            )
        if not kinun_candidates:
            kinun_candidates = sorted(data_dir.glob(disco.kinun_glob))

    if not plan_files:
        raise FileNotFoundError(
            f"No plan files found in {input_dir} (or {data_dir if data_dir.exists() else 'data/'}) "
            f"with pattern: {disco.plan_glob}"
        )

    kinun_file: Optional[Path] = None
    if kinun_candidates:
        kinun_file = _pick_newest(kinun_candidates)
    else:
        log.info(
            "No kinun Excel found in %s (or %s) with pattern %s. "
            "Continuing without it (JSON baseline is expected).",
            input_dir,
            data_dir if data_dir.exists() else (input_dir / "data"),
            disco.kinun_glob,
        )

    return plan_files, kinun_file

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import ast

RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

def apply_red_highlights(output_xlsx_path: str, all_checks_df: pd.DataFrame) -> None:
    """
    all_checks_df is the All_Checks dataframe (exported from CheckResult.to_record()).
    It must include a column named 'excel_cells' which can be:
      - None/NaN
      - list[str]
      - string representation of list[str] (depends how pandas wrote it)
    """
    if all_checks_df is None or len(all_checks_df) == 0:
        return
    if "excel_cells" not in all_checks_df.columns:
        return

    wb = load_workbook(output_xlsx_path)

    for v in all_checks_df["excel_cells"].dropna():
        cells = None
        if isinstance(v, list):
            cells = v
        else:
            s = str(v).strip()
            if not s:
                continue
            # pandas often stores lists as string like "['Sheet!B12']"
            try:
                parsed = ast.literal_eval(s)
                if isinstance(parsed, list):
                    cells = parsed
            except Exception:
                # fallback: single "Sheet!B12"
                cells = [s] if "!" in s else None

        if not cells:
            continue

        for ref in cells:
            if not isinstance(ref, str) or "!" not in ref:
                continue
            sh, addr = ref.split("!", 1)
            if sh not in wb.sheetnames:
                continue
            ws = wb[sh]
            try:
                ws[addr].fill = RED_FILL
            except Exception:
                # ignore invalid addresses
                pass

    wb.save(output_xlsx_path)



def ensure_sheet_exists(xlsx_path: Path, sheet_name: str) -> None:
    xl = pd.ExcelFile(xlsx_path, engine="openpyxl")
    if sheet_name not in xl.sheet_names:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {xlsx_path.name}. "
            f"Available: {xl.sheet_names}"
        )


def load_plan_sheet_with_header_fix(plan_file: str | Path, cfg: PlanConfig) -> tuple[pd.DataFrame, PlanLoadInfo]:
    plan_file = Path(plan_file)
    ensure_sheet_exists(plan_file, cfg.sheet_name)

    raw = pd.read_excel(plan_file, sheet_name=cfg.sheet_name, header=None, engine="openpyxl")

    # Find data start row: first 'מים' in column B
    data_start_idx = None
    for i in range(len(raw)):
        v = normalize_text(raw.iat[i, cfg.data_marker_col_idx] if cfg.data_marker_col_idx < raw.shape[1] else None)
        if v == "מים":
            data_start_idx = i
            break
    if data_start_idx is None:
        raise ValueError(f"[{plan_file.name}] Could not detect data start row (expected 'מים' in column B).")

    data_start_excel_row = data_start_idx + 1  # 1-based
    cfg.data_start_excel_row = data_start_excel_row  # anchor for excel_row_to_df_index()

    # Pick best header row from window above data_start_idx
    lookback = max(1, int(cfg.header_lookback_rows))
    candidates = range(max(0, data_start_idx - lookback), data_start_idx)

    best_header_idx = None
    best_score = -1
    for r in candidates:
        row_vals = raw.iloc[r].tolist()
        score = sum(1 for x in row_vals if is_city_like_header(x))
        if score > best_score:
            best_score = score
            best_header_idx = r

    if best_header_idx is None or best_score <= 0:
        raise ValueError(
            f"[{plan_file.name}] Could not detect a city header row "
            f"(no city-like headers in {lookback} rows above first 'מים')."
        )

    headers = [normalize_text(x) for x in raw.iloc[best_header_idx].tolist()]

    df = raw.iloc[data_start_idx:].copy()
    df.columns = headers
    df.reset_index(drop=True, inplace=True)

    info = PlanLoadInfo(
        data_start_idx_0based=data_start_idx,
        header_row_idx_0based=best_header_idx,
        data_start_excel_row_1based=data_start_excel_row,
    )
    log.info("Loaded plan %s (data_start_excel_row=%s, header_row=%s, header_score=%s)",
             plan_file.name, data_start_excel_row, best_header_idx + 1, best_score)

    return df, info


def load_kinun_reference(kinun_file: str | Path) -> pd.DataFrame:
    kinun_file = Path(kinun_file)
    raw = pd.read_excel(kinun_file, header=None, usecols="A:E", engine="openpyxl")

    def is_data_row(row: pd.Series) -> bool:
        a = normalize_text(row.iloc[0])
        b, c, d, e = row.iloc[1], row.iloc[2], row.iloc[3], row.iloc[4]
        if not a:
            return False
        if a in {"תאגיד מים וביוב", "תאגיד", "תאגיד מים"}:
            return False
        return any(pd.notna(x) and isinstance(x, (int, float)) for x in [b, c, d, e])

    start_idx = None
    for i in range(len(raw)):
        if is_data_row(raw.iloc[i]):
            start_idx = i
            break
    if start_idx is None:
        raise ValueError(f"[{kinun_file.name}] Could not detect first data row in kinun file (A–E).")

    df = raw.iloc[start_idx:].copy()
    df.columns = [
        "תאגיד מים וביוב",
        "תשתיות מים מלא",
        "תשתיות מים מופחת",
        "תשתיות ביוב מלא",
        "תשתיות ביוב מופחת",
    ]
    df = df[df["תאגיד מים וביוב"].notna()].copy()
    df["תאגיד מים וביוב"] = df["תאגיד מים וביוב"].map(normalize_text)
    return df



import pandas as pd
import re
def _norm_header_cell(x: object) -> str:
    # pandas uses NaN for empty header cells; str(NaN) == "nan"
    try:
        import pandas as pd
        if x is None or pd.isna(x):
            return ""
    except Exception:
        if x is None:
            return ""

    s = str(x)
    s = s.replace("\u00A0", " ").replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = " ".join(s.split()).strip()
    return s
import re
import pandas as pd

def _clean_report_header(df: pd.DataFrame) -> pd.DataFrame:
    """
    Fix 'גיליון דיווח' headers that come as MultiIndex (2-row header),
    and remove helper fragments like 'רחוב/שכונה/תב"ע' and 'Unnamed: ...'.
    Result: flat, clean, stable column names like 'שם פרויקט', 'מיקום פרויקט', 'סיווג פרויקט'.
    """

    LOCATION_HELPER = {'רחוב/שכונה/תב"ע', 'רחוב/שכונה/תב״ע', 'רחוב/שכונה/תב”ע'}

    def _norm_piece(x: object) -> str:
        s = "" if x is None else str(x)
        s = s.replace("\u00A0", " ")
        s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")
        s = " ".join(s.split()).strip()
        return s

    if isinstance(df.columns, pd.MultiIndex):
        new_cols = []
        for tup in df.columns.to_list():
            parts = [_norm_piece(p) for p in tup if _norm_piece(p)]
            # drop helper fragments + Unnamed
            parts = [p for p in parts if p not in LOCATION_HELPER and not p.startswith("Unnamed:")]
            col = " ".join(parts).strip()
            # if everything dropped, keep original first non-empty (fallback)
            if not col:
                col = _norm_piece(tup[0]) if len(tup) > 0 else ""
            new_cols.append(col)

        df = df.copy()
        df.columns = new_cols
        return df

    # non-multiindex: still clean Unnamed/helper fragments if they appear in string columns
    cleaned = []
    for c in df.columns:
        s = _norm_piece(c)
        for h in LOCATION_HELPER:
            s = s.replace(h, "")
        s = " ".join(s.split()).strip()
        cleaned.append(s)

    df = df.copy()
    df.columns = cleaned
    return df



def load_report_sheet(xlsx_path: str, sheet_name: str, header_row: int) -> pd.DataFrame:
    """
    Loads 'report' sheet (e.g., "גיליון דיווח") with a configurable header row.
    Supports 2-row headers (merged cells) by flattening MultiIndex columns.
    header_row is 0-based for pandas (Excel row 7 => header_row=6).
    """
    df = pd.read_excel(
        xlsx_path,
        sheet_name=sheet_name,
        header=[header_row, header_row + 1],  # ✅ 2-row header
        engine="openpyxl",
    )
    df = _clean_report_header(df)  # ✅ flatten & clean
    return df
